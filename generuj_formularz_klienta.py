#!/usr/bin/env python3
"""
Generator formularza zbierania danych od klienta (intake form).
Tworzy profesjonalny XLSX z checklistą dokumentów i informacji.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os


def create_intake_form():
    wb = Workbook()

    # Style
    header_font = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
    header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    section_font = Font(name='Calibri', bold=True, size=11, color='003366')
    section_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    label_font = Font(name='Calibri', size=10)
    input_fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
    check_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    note_font = Font(name='Calibri', size=9, italic=True, color='666666')

    def set_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def add_header(ws, row, text, cols=5):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 35

    def add_section(ws, row, text, cols=5):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = section_font
        cell.fill = section_fill
        ws.row_dimensions[row].height = 25

    def add_field(ws, row, label, col_span=2, input_cols=None, note=''):
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = label_font
        c1.border = thin_border
        c1.alignment = Alignment(vertical='center', wrap_text=True)

        if input_cols is None:
            input_cols = [3, 4, 5]
        for col in input_cols:
            c = ws.cell(row=row, column=col)
            c.fill = input_fill
            c.border = thin_border

        if note:
            c_note = ws.cell(row=row, column=max(input_cols) + 1 if max(input_cols) < 6 else 5, value=note)
            c_note.font = note_font

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)

    def add_checklist_row(ws, row, item, status_col=3, note_col=4):
        c1 = ws.cell(row=row, column=1, value=item)
        c1.font = label_font
        c1.border = thin_border
        c1.alignment = Alignment(vertical='center', wrap_text=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

        c2 = ws.cell(row=row, column=status_col)
        c2.fill = check_fill
        c2.border = thin_border
        c2.alignment = Alignment(horizontal='center')

        for col in range(status_col + 1, 6):
            c = ws.cell(row=row, column=col)
            c.fill = input_fill
            c.border = thin_border

    # ========================================
    # ARKUSZ 1: DANE KLIENTA
    # ========================================
    ws1 = wb.active
    ws1.title = 'Dane klienta'
    set_widths(ws1, [35, 15, 20, 20, 20])
    ws1.sheet_properties.tabColor = '003366'

    row = 1
    add_header(ws1, row, 'FORMULARZ ZBIERANIA DANYCH OD KLIENTA')
    row += 2

    # Dane podstawowe
    add_section(ws1, row, 'A. DANE PODSTAWOWE KLIENTA')
    row += 1
    fields_basic = [
        'Nazwa firmy (pełna)',
        'NIP',
        'Adres siedziby',
        'Adres korespondencyjny (jeśli inny)',
        'Osoba kontaktowa (imię, nazwisko)',
        'Telefon',
        'E-mail',
        'Branża / profil działalności',
        'Forma prawna (sp. z o.o., S.A., JDG, itp.)',
        'Liczba pracowników (orientacyjna)',
    ]
    for f in fields_basic:
        add_field(ws1, row, f)
        row += 1

    row += 1
    add_section(ws1, row, 'B. PROFIL DZIAŁALNOŚCI')
    row += 1
    fields_profile = [
        ('Dni pracy w tygodniu', 'np. Pn-Pt / Pn-Sob / 7 dni'),
        ('Godziny pracy (zmiany)', 'np. 6:00-22:00 (2 zmiany) / 24h'),
        ('Sezonowość produkcji', 'np. wyższa latem / równomierna / zimą'),
        ('Główne odbiorniki energii', 'np. linie produkcyjne, chłodnie, sprężarki'),
        ('Planowane zmiany w zużyciu', 'np. nowa linia produkcyjna, rozbudowa'),
        ('Czy produkcja jest ciągła?', 'TAK / NIE'),
    ]
    for label, note in fields_profile:
        add_field(ws1, row, label, note=note)
        row += 1

    row += 1
    add_section(ws1, row, 'C. PUNKTY POBORU ENERGII (PPE)')
    row += 1
    # Nagłówki tabeli PPE
    ppe_headers = ['Nr PPE', 'Adres PPE', 'Moc umowna (kW)', 'Moc przyłącz. (kW)', 'Grupa taryfowa']
    for col, h in enumerate(ppe_headers, 1):
        cell = ws1.cell(row=row, column=col, value=h)
        cell.font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    row += 1
    # 5 pustych wierszy na PPE
    for _ in range(5):
        for col in range(1, 6):
            cell = ws1.cell(row=row, column=col)
            cell.fill = input_fill
            cell.border = thin_border
        row += 1

    row += 1
    add_section(ws1, row, 'D. PUNKTY POBORU GAZU (PPG)')
    row += 1
    ppg_headers = ['Nr PPG', 'Adres PPG', 'Moc umowna (kWh/h)', 'Grupa taryfowa', 'Cel zużycia gazu']
    for col, h in enumerate(ppg_headers, 1):
        cell = ws1.cell(row=row, column=col, value=h)
        cell.font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    row += 1
    for _ in range(3):
        for col in range(1, 6):
            cell = ws1.cell(row=row, column=col)
            cell.fill = input_fill
            cell.border = thin_border
        row += 1

    # ========================================
    # ARKUSZ 2: CHECKLIST DOKUMENTÓW
    # ========================================
    ws2 = wb.create_sheet('Checklist dokumentów')
    set_widths(ws2, [40, 10, 12, 15, 25])
    ws2.sheet_properties.tabColor = '006600'

    row = 1
    add_header(ws2, row, 'CHECKLIST DOKUMENTÓW DO ZEBRANIA')
    row += 2

    # Nagłówki
    col_headers = ['Dokument', 'Dotyczy', 'Otrzymano?', 'Data', 'Uwagi']
    for col, h in enumerate(col_headers, 1):
        cell = ws2.cell(row=row, column=col, value=h)
        cell.font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
        cell.fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    row += 1

    # Dropdown validation TAK/NIE
    dv_yn = DataValidation(type='list', formula1='"TAK,NIE,W TOKU"', allow_blank=True)
    ws2.add_data_validation(dv_yn)

    # Dokumenty
    add_section(ws2, row, 'ENERGIA ELEKTRYCZNA')
    row += 1

    ee_docs = [
        ('Aktualna umowa na energię elektryczną', 'ee'),
        ('FV za ee – miesiąc 1 (najstarszy)', 'PPE 1'),
        ('FV za ee – miesiąc 2', 'PPE 1'),
        ('FV za ee – miesiąc 3', 'PPE 1'),
        ('FV za ee – miesiąc 4', 'PPE 1'),
        ('FV za ee – miesiąc 5', 'PPE 1'),
        ('FV za ee – miesiąc 6', 'PPE 1'),
        ('FV za ee – miesiąc 7', 'PPE 1'),
        ('FV za ee – miesiąc 8', 'PPE 1'),
        ('FV za ee – miesiąc 9', 'PPE 1'),
        ('FV za ee – miesiąc 10', 'PPE 1'),
        ('FV za ee – miesiąc 11', 'PPE 1'),
        ('FV za ee – miesiąc 12 (najnowszy)', 'PPE 1'),
        ('--- Kolejne PPE: powtórzyć 12 FV ---', ''),
    ]
    for doc, scope in ee_docs:
        c1 = ws2.cell(row=row, column=1, value=doc)
        c1.font = label_font
        c1.border = thin_border
        c2 = ws2.cell(row=row, column=2, value=scope)
        c2.font = label_font
        c2.border = thin_border
        c2.alignment = Alignment(horizontal='center')
        c3 = ws2.cell(row=row, column=3)
        c3.fill = check_fill
        c3.border = thin_border
        dv_yn.add(c3)
        for col in [4, 5]:
            c = ws2.cell(row=row, column=col)
            c.fill = input_fill
            c.border = thin_border
        row += 1

    row += 1
    add_section(ws2, row, 'GAZ ZIEMNY')
    row += 1

    gas_docs = [
        ('Aktualna umowa na gaz ziemny', 'gaz'),
        ('FV za gaz – miesiąc 1 (najstarszy)', 'PPG 1'),
        ('FV za gaz – miesiąc 2', 'PPG 1'),
        ('FV za gaz – miesiąc 3', 'PPG 1'),
        ('FV za gaz – miesiąc 4', 'PPG 1'),
        ('FV za gaz – miesiąc 5', 'PPG 1'),
        ('FV za gaz – miesiąc 6', 'PPG 1'),
        ('FV za gaz – miesiąc 7', 'PPG 1'),
        ('FV za gaz – miesiąc 8', 'PPG 1'),
        ('FV za gaz – miesiąc 9', 'PPG 1'),
        ('FV za gaz – miesiąc 10', 'PPG 1'),
        ('FV za gaz – miesiąc 11', 'PPG 1'),
        ('FV za gaz – miesiąc 12 (najnowszy)', 'PPG 1'),
    ]
    for doc, scope in gas_docs:
        c1 = ws2.cell(row=row, column=1, value=doc)
        c1.font = label_font
        c1.border = thin_border
        c2 = ws2.cell(row=row, column=2, value=scope)
        c2.font = label_font
        c2.border = thin_border
        c2.alignment = Alignment(horizontal='center')
        c3 = ws2.cell(row=row, column=3)
        c3.fill = check_fill
        c3.border = thin_border
        dv_yn.add(c3)
        for col in [4, 5]:
            c = ws2.cell(row=row, column=col)
            c.fill = input_fill
            c.border = thin_border
        row += 1

    row += 1
    add_section(ws2, row, 'DODATKOWE DOKUMENTY')
    row += 1

    extra_docs = [
        ('Umowa dystrybucyjna (OSD)', 'ee'),
        ('Warunki przyłączenia', 'ee'),
        ('Schemat elektryczny zakładu', 'techniczny'),
        ('Dane z licznika 15-min (jeśli dostępne)', 'ee'),
        ('Dokumentacja istniejącej instalacji PV', 'OZE'),
        ('Dokumentacja pompy ciepła / kogeneracji', 'OZE'),
        ('Wyniki audytu energetycznego (jeśli był)', 'audyt'),
        ('Mapka/plan zakładu (dla PV na dachu)', 'PV'),
    ]
    for doc, scope in extra_docs:
        c1 = ws2.cell(row=row, column=1, value=doc)
        c1.font = label_font
        c1.border = thin_border
        c2 = ws2.cell(row=row, column=2, value=scope)
        c2.font = label_font
        c2.border = thin_border
        c2.alignment = Alignment(horizontal='center')
        c3 = ws2.cell(row=row, column=3)
        c3.fill = check_fill
        c3.border = thin_border
        dv_yn.add(c3)
        for col in [4, 5]:
            c = ws2.cell(row=row, column=col)
            c.fill = input_fill
            c.border = thin_border
        row += 1

    # ========================================
    # ARKUSZ 3: INFORMACJE TECHNICZNE
    # ========================================
    ws3 = wb.create_sheet('Informacje techniczne')
    set_widths(ws3, [45, 10, 20, 20, 20])
    ws3.sheet_properties.tabColor = 'CC6600'

    row = 1
    add_header(ws3, row, 'INFORMACJE TECHNICZNE DO ZEBRANIA OD KLIENTA')
    row += 2

    # Dropdown TAK/NIE
    dv_yn3 = DataValidation(type='list', formula1='"TAK,NIE,N/D"', allow_blank=True)
    ws3.add_data_validation(dv_yn3)

    add_section(ws3, row, 'E. ISTNIEJĄCE INSTALACJE OZE')
    row += 1

    oze_questions = [
        ('Czy jest zainstalowana fotowoltaika (PV)?', 'Jeśli TAK → moc kWp, rok instalacji, producent'),
        ('Moc instalacji PV (kWp)', ''),
        ('Rok instalacji PV', ''),
        ('Producent falownika PV', ''),
        ('Roczna produkcja PV (kWh/rok)', 'Z danych z falownika lub FV'),
        ('Obecna autokonsumpcja PV (%)', 'Ile % produkcji PV zużywa zakład bezpośrednio'),
        ('Czy nadwyżki idą do sieci?', 'Net-billing / net-metering / sprzedaż'),
        ('Czy jest pompa ciepła?', 'Jeśli TAK → typ, moc, rok'),
        ('Czy jest kogeneracja (CHP)?', 'Jeśli TAK → typ, moc el./cieplna'),
        ('Czy jest magazyn energii (BESS)?', 'Jeśli TAK → pojemność kWh, producent'),
        ('Czy jest magazyn ciepła?', 'Jeśli TAK → pojemność, typ'),
    ]
    for q, note in oze_questions:
        c1 = ws3.cell(row=row, column=1, value=q)
        c1.font = label_font
        c1.border = thin_border
        c1.alignment = Alignment(wrap_text=True)

        c2 = ws3.cell(row=row, column=2)
        c2.fill = check_fill
        c2.border = thin_border
        if 'Czy' in q:
            dv_yn3.add(c2)

        ws3.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        c3 = ws3.cell(row=row, column=3)
        c3.fill = input_fill
        c3.border = thin_border
        ws3.cell(row=row, column=4).border = thin_border

        c5 = ws3.cell(row=row, column=5, value=note)
        c5.font = note_font
        c5.border = thin_border
        c5.alignment = Alignment(wrap_text=True)
        row += 1

    row += 1
    add_section(ws3, row, 'F. INFRASTRUKTURA ELEKTRYCZNA')
    row += 1

    infra_questions = [
        ('Czy jest wykonana kompensacja mocy biernej?', 'Bateria kondensatorów / KMB'),
        ('Typ kompensacji (jeśli jest)', 'Stała / automatyczna / aktywna'),
        ('Moc kompensacji (kvar)', ''),
        ('Czy jest zainstalowany agregat prądotwórczy?', ''),
        ('Moc agregatu (kVA)', ''),
        ('Czy jest UPS?', 'Jeśli TAK → moc, pojemność'),
        ('Czy są problemy z jakością energii?', 'Wahania napięcia, harmoniczne, spadki'),
        ('Czy jest stacja transformatorowa własna?', ''),
        ('Napięcie zasilania (nn/SN)', 'np. 400V, 15kV, 20kV'),
        ('Czy jest wolna przestrzeń na magazyn energii?', 'Wewnątrz / na zewnątrz / dach'),
        ('Czy jest wolna przestrzeń na dodatkowe PV?', 'Dach / grunt / wiata'),
        ('Powierzchnia dostępna na PV (m²)', 'Orientacyjna'),
    ]
    for q, note in infra_questions:
        c1 = ws3.cell(row=row, column=1, value=q)
        c1.font = label_font
        c1.border = thin_border
        c1.alignment = Alignment(wrap_text=True)

        c2 = ws3.cell(row=row, column=2)
        c2.fill = check_fill
        c2.border = thin_border
        if 'Czy' in q:
            dv_yn3.add(c2)

        ws3.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        c3 = ws3.cell(row=row, column=3)
        c3.fill = input_fill
        c3.border = thin_border
        ws3.cell(row=row, column=4).border = thin_border

        c5 = ws3.cell(row=row, column=5, value=note)
        c5.font = note_font
        c5.border = thin_border
        c5.alignment = Alignment(wrap_text=True)
        row += 1

    row += 1
    add_section(ws3, row, 'G. OCZEKIWANIA KLIENTA')
    row += 1

    expect_questions = [
        ('Główny cel (oszczędność kosztów / niezależność / ESG)', ''),
        ('Czy są potrzebne gwarancje pochodzenia (zielona energia)?', 'GO / brak'),
        ('Preferowany model zakupu energii', 'FIX / RDN / MIX / brak preferencji'),
        ('Oczekiwany budżet inwestycyjny (orientacyjny)', ''),
        ('Preferowany model finansowania', 'Zakup / leasing / ESCO / PPA / raty'),
        ('Planowany termin realizacji', ''),
        ('Czy firma uczestniczy w programach ESG/CSR?', ''),
        ('Czy firma raportuje emisje CO₂?', ''),
        ('Czy firma jest zainteresowana DSR (redukcja popytu)?', ''),
        ('Inne uwagi / oczekiwania klienta', ''),
    ]
    for q, note in expect_questions:
        c1 = ws3.cell(row=row, column=1, value=q)
        c1.font = label_font
        c1.border = thin_border
        c1.alignment = Alignment(wrap_text=True)

        c2 = ws3.cell(row=row, column=2)
        c2.fill = check_fill
        c2.border = thin_border

        ws3.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        c3 = ws3.cell(row=row, column=3)
        c3.fill = input_fill
        c3.border = thin_border
        ws3.cell(row=row, column=4).border = thin_border

        c5 = ws3.cell(row=row, column=5, value=note)
        c5.font = note_font
        c5.border = thin_border
        c5.alignment = Alignment(wrap_text=True)
        row += 1

    # ========================================
    # ARKUSZ 4: ANALIZA UMOWY
    # ========================================
    ws4 = wb.create_sheet('Analiza umowy')
    set_widths(ws4, [40, 15, 20, 20, 20])
    ws4.sheet_properties.tabColor = 'CC0000'

    row = 1
    add_header(ws4, row, 'ANALIZA AKTUALNEJ UMOWY NA ENERGIĘ / GAZ')
    row += 2

    add_section(ws4, row, 'H. UMOWA NA ENERGIĘ ELEKTRYCZNĄ')
    row += 1

    contract_fields = [
        ('Sprzedawca energii', ''),
        ('Nr umowy', ''),
        ('Data zawarcia umowy', ''),
        ('Data końca umowy', '← KLUCZOWE: od kiedy można podpisać nową'),
        ('Okres wypowiedzenia', 'np. 1 miesiąc, 3 miesiące'),
        ('Czy jest auto-prolongata?', 'TAK/NIE + warunki'),
        ('Rodzaj umowy (kompleksowa / rozdzielona)', ''),
        ('Cena energii w umowie (PLN/kWh netto)', ''),
        ('Czy cena jest stała (FIX) czy zmienna?', ''),
        ('Opłata handlowa (PLN/mies.)', ''),
        ('Operator Sieci Dystrybucyjnej (OSD)', 'Tauron/Enea/Energa/PGE/innogy'),
        ('Grupa taryfowa', 'np. C21, C22a, C22b, B21, B23'),
        ('Moc umowna (kW)', ''),
        ('Moc przyłączeniowa (kW)', ''),
        ('Kary za przekroczenie mocy (czy występowały)', ''),
        ('Opłata mocowa – kategoria (K1-K4)', ''),
        ('Czy jest klauzula waloryzacyjna?', ''),
        ('Czy są kary za wcześniejsze rozwiązanie?', ''),
    ]
    for label, note in contract_fields:
        c1 = ws4.cell(row=row, column=1, value=label)
        c1.font = label_font
        c1.border = thin_border
        c1.alignment = Alignment(wrap_text=True)

        ws4.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        c2 = ws4.cell(row=row, column=2)
        c2.fill = input_fill
        c2.border = thin_border
        ws4.cell(row=row, column=3).border = thin_border

        ws4.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
        c4 = ws4.cell(row=row, column=4, value=note)
        c4.font = note_font
        c4.border = thin_border
        c4.alignment = Alignment(wrap_text=True)
        ws4.cell(row=row, column=5).border = thin_border
        row += 1

    row += 1
    add_section(ws4, row, 'I. UMOWA NA GAZ ZIEMNY')
    row += 1

    gas_fields = [
        ('Sprzedawca gazu', ''),
        ('Nr umowy', ''),
        ('Data zawarcia umowy', ''),
        ('Data końca umowy', '← od kiedy można podpisać nową'),
        ('Okres wypowiedzenia', ''),
        ('Cena gazu w umowie (PLN/kWh netto)', ''),
        ('Czy cena jest stała (FIX) czy zmienna?', ''),
        ('Roczne zużycie gazu (kWh lub m³)', ''),
        ('Cel zużycia gazu', 'Ogrzewanie / proces / CHP'),
    ]
    for label, note in gas_fields:
        c1 = ws4.cell(row=row, column=1, value=label)
        c1.font = label_font
        c1.border = thin_border
        c1.alignment = Alignment(wrap_text=True)

        ws4.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        c2 = ws4.cell(row=row, column=2)
        c2.fill = input_fill
        c2.border = thin_border
        ws4.cell(row=row, column=3).border = thin_border

        ws4.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
        c4 = ws4.cell(row=row, column=4, value=note)
        c4.font = note_font
        c4.border = thin_border
        c4.alignment = Alignment(wrap_text=True)
        ws4.cell(row=row, column=5).border = thin_border
        row += 1

    # ========================================
    # ARKUSZ 5: DANE Z FAKTUR (szablon)
    # ========================================
    ws5 = wb.create_sheet('Dane z faktur ee')
    set_widths(ws5, [15, 15, 15, 15, 15, 15, 15, 15, 15, 15])
    ws5.sheet_properties.tabColor = '4472C4'

    row = 1
    add_header(ws5, row, 'DANE Z FAKTUR ZA ENERGIĘ ELEKTRYCZNĄ (12 miesięcy)', cols=10)
    row += 2

    ws5.cell(row=row, column=1, value='PPE nr:').font = section_font
    ws5.cell(row=row, column=2).fill = input_fill
    ws5.cell(row=row, column=2).border = thin_border
    row += 2

    fv_headers = [
        'Miesiąc', 'Zużycie\n(kWh)', 'Koszt energii\n(PLN netto)',
        'Koszt dystr.\n(PLN netto)', 'Opł. mocowa\n(PLN netto)',
        'Opł. OZE\n(PLN netto)', 'Opł. kogener.\n(PLN netto)',
        'Akcyza\n(PLN)', 'RAZEM netto\n(PLN)', 'Moc max.\n(kW)'
    ]
    for col, h in enumerate(fv_headers, 1):
        cell = ws5.cell(row=row, column=col, value=h)
        cell.font = Font(name='Calibri', bold=True, size=9, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')
    ws5.row_dimensions[row].height = 40
    row += 1

    months = ['Styczeń', 'Luty', 'Marzec', 'Kwiecień', 'Maj', 'Czerwiec',
              'Lipiec', 'Sierpień', 'Wrzesień', 'Październik', 'Listopad', 'Grudzień']
    for m in months:
        ws5.cell(row=row, column=1, value=m).font = label_font
        ws5.cell(row=row, column=1).border = thin_border
        for col in range(2, 11):
            c = ws5.cell(row=row, column=col)
            c.fill = input_fill
            c.border = thin_border
            c.number_format = '#,##0.00' if col != 2 else '#,##0'
        row += 1

    # Wiersz SUMA
    ws5.cell(row=row, column=1, value='SUMA / ŚREDNIA').font = Font(name='Calibri', bold=True, size=10)
    ws5.cell(row=row, column=1).border = thin_border
    for col in range(2, 11):
        c = ws5.cell(row=row, column=col)
        c.border = thin_border
        c.font = Font(name='Calibri', bold=True, size=10)
        # Formuła SUM
        col_letter = get_column_letter(col)
        start = row - 12
        end = row - 1
        if col == 10:  # Moc max = MAX
            c.value = f'=MAX({col_letter}{start}:{col_letter}{end})'
        else:
            c.value = f'=SUM({col_letter}{start}:{col_letter}{end})'
        c.number_format = '#,##0'

    # ========================================
    # ARKUSZ 6: WORKFLOW
    # ========================================
    ws6 = wb.create_sheet('Workflow analizy')
    set_widths(ws6, [8, 45, 15, 15, 25])
    ws6.sheet_properties.tabColor = '7030A0'

    row = 1
    add_header(ws6, row, 'WORKFLOW ANALIZY I PRZYGOTOWANIA OFERTY')
    row += 2

    wf_headers = ['Krok', 'Zadanie', 'Odpow.', 'Status', 'Uwagi']
    for col, h in enumerate(wf_headers, 1):
        cell = ws6.cell(row=row, column=col, value=h)
        cell.font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
        cell.fill = PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid')
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    row += 1

    dv_status = DataValidation(type='list', formula1='"Do zrobienia,W toku,Gotowe,N/D"', allow_blank=True)
    ws6.add_data_validation(dv_status)

    workflow_steps = [
        # Faza 1: Zbieranie danych
        ('', 'FAZA 1: ZBIERANIE DANYCH', '', '', ''),
        ('1.1', 'Zebranie danych klienta (arkusz "Dane klienta")', '', '', ''),
        ('1.2', 'Zebranie 12 FV za ee dla każdego PPE', '', '', ''),
        ('1.3', 'Zebranie 12 FV za gaz dla każdego PPG', '', '', ''),
        ('1.4', 'Zebranie aktualnych umów ee i gaz', '', '', ''),
        ('1.5', 'Uzupełnienie informacji technicznych', '', '', ''),
        ('1.6', 'Wprowadzenie danych z FV do arkusza', '', '', ''),
        # Faza 2: Analiza
        ('', 'FAZA 2: ANALIZA', '', '', ''),
        ('2.1', 'Analiza umów – termin zakończenia, warunki', '', '', ''),
        ('2.2', 'Potwierdzenie rezerwacji i warunków płatności', '', '', ''),
        ('2.3', 'Ustalenie od kiedy można podpisać nową umowę', '', '', ''),
        ('2.4', 'Analiza profilu zużycia (wolumen, sezonowość)', '', '', ''),
        ('2.5', 'Analiza kosztów dystrybucji (30% od oszczędności)', '', '', ''),
        ('2.6', 'Rekomendacja produktu: FIX / RDN / MIX + wycena', '', '', ''),
        ('2.7', 'Analiza opłaty mocowej i potencjału peak shaving', '', '', ''),
        ('2.8', 'Analiza kompensacji mocy biernej (KMB)', '', '', ''),
        # Faza 3: Rekomendacje OZE
        ('', 'FAZA 3: REKOMENDACJE PRODUKTÓW DODATKOWYCH', '', '', ''),
        ('3.1', 'Rekomendacja PV – sizing na bazie profilu (SUN HELP)', '', '', ''),
        ('3.2', 'Rekomendacja BESS – sizing + arbitraż + rynek mocy (ALIANS)', '', '', 'Może być stand-alone'),
        ('3.3', 'Rekomendacja DSR (ALIANS)', '', '', ''),
        ('3.4', 'Rekomendacja KMB – kompensacja mocy biernej', '', '', ''),
        # Faza 4: Oferta
        ('', 'FAZA 4: PRZYGOTOWANIE OFERTY', '', '', ''),
        ('4.1', 'Kalkulacja ROI dla każdego produktu', '', '', ''),
        ('4.2', 'Przygotowanie opcji finansowania', '', '', 'Leasing/ESCO/PPA/raty'),
        ('4.3', 'Generowanie oferty XLSX z kalkulatora', '', '', ''),
        ('4.4', 'Przygotowanie prezentacji dla klienta', '', '', ''),
        ('4.5', 'Review wewnętrzny oferty', '', '', ''),
        # Faza 5: Prezentacja
        ('', 'FAZA 5: PREZENTACJA I FOLLOW-UP', '', '', ''),
        ('5.1', 'Spotkanie z klientem – prezentacja oferty', '', '', ''),
        ('5.2', 'Follow-up – odpowiedzi na pytania', '', '', ''),
        ('5.3', 'Negocjacje warunków', '', '', ''),
        ('5.4', 'Podpisanie umowy', '', '', ''),
    ]

    for step, task, resp, status, notes in workflow_steps:
        if not step:
            # Section header
            add_section(ws6, row, task)
            row += 1
            continue

        c1 = ws6.cell(row=row, column=1, value=step)
        c1.font = label_font
        c1.border = thin_border
        c1.alignment = Alignment(horizontal='center')

        c2 = ws6.cell(row=row, column=2, value=task)
        c2.font = label_font
        c2.border = thin_border
        c2.alignment = Alignment(wrap_text=True)

        c3 = ws6.cell(row=row, column=3)
        c3.fill = input_fill
        c3.border = thin_border

        c4 = ws6.cell(row=row, column=4)
        c4.fill = check_fill
        c4.border = thin_border
        dv_status.add(c4)

        c5 = ws6.cell(row=row, column=5, value=notes)
        c5.font = note_font
        c5.border = thin_border
        c5.alignment = Alignment(wrap_text=True)

        row += 1

    # Zapisz
    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                               'Formularz_Dane_Klienta.xlsx')
    wb.save(output_path)
    print(f'Formularz zapisany: {output_path}')
    return output_path


def create_intake_form_bytes() -> bytes:
    """Generuje formularz XLSX i zwraca jako bytes (do st.download_button)."""
    path = create_intake_form()
    with open(path, 'rb') as f:
        return f.read()


if __name__ == '__main__':
    create_intake_form()
