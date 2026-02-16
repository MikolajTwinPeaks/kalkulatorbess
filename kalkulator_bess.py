#!/usr/bin/env python3
"""
Kalkulator BESS (Battery Energy Storage System) dla zakładów produkcyjnych z PV.

Wprowadź dane z faktury za energię elektryczną, a kalkulator:
- Oszacuje optymalną pojemność magazynu
- Wyliczy oszczędności z arbitrażu cenowego, peak shavingu i autokonsumpcji PV
- Wygeneruje ofertę w formacie XLSX

Użycie:
    python3 kalkulator_bess.py
    (program zapyta o dane interaktywnie)

    python3 kalkulator_bess.py --demo
    (uruchomi z danymi przykładowymi)
"""

import sys
import os
from dataclasses import dataclass
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter


@dataclass
class DaneFaktury:
    """Dane z faktury za energię elektryczną."""
    nazwa_firmy: str
    roczne_zuzycie_kwh: float          # kWh/rok
    moc_zamowiona_kw: float             # kW
    moc_pv_kwp: float                   # kWp zainstalowanego PV
    roczna_produkcja_pv_kwh: float      # kWh/rok produkcji PV
    autokonsumpcja_pv_procent: float    # % autokonsumpcji (bez magazynu)
    cena_energii_pln_kwh: float         # PLN/kWh (średnia cena z faktury)
    oplata_dystrybucyjna_pln_kwh: float # PLN/kWh
    oplata_mocowa_pln_mwh: float        # PLN/MWh (stawka opłaty mocowej)
    kategoria_mocowa: str               # K1-K4
    sredni_rachunek_miesiac_pln: float  # PLN/mies. (średni rachunek)


@dataclass
class ParametryBESS:
    """Parametry systemu BESS."""
    pojemnosc_kwh: float
    moc_kw: float
    rte: float                  # Round-Trip Efficiency (0-1)
    cykle_dzienne: float
    koszt_pln_kwh: float        # PLN/kWh CAPEX
    zywotnosc_lat: int
    degradacja_roczna: float    # % rocznie


@dataclass
class WynikiKalkulacji:
    """Wyniki kalkulacji oszczędności."""
    # CAPEX
    capex_bess_pln: float
    capex_ems_pln: float
    capex_instalacja_pln: float
    capex_calkowity_pln: float

    # Oszczędności roczne
    oszczednosc_autokonsumpcja_pln: float
    oszczednosc_arbitraz_pln: float
    oszczednosc_peak_shaving_pln: float
    oszczednosc_calkowita_roczna_pln: float

    # Zwrot
    okres_zwrotu_lat: float
    roi_10lat_procent: float
    npv_10lat_pln: float

    # Dane dodatkowe
    nowa_autokonsumpcja_procent: float
    redukcja_rachunku_procent: float
    oszczednosc_co2_ton_rok: float


def oblicz_parametry_bess(dane: DaneFaktury) -> ParametryBESS:
    """Dobiera optymalną pojemność magazynu na podstawie danych z faktury."""

    # Nadwyżka PV dziennie (średnia roczna)
    nadwyzka_pv_roczna = dane.roczna_produkcja_pv_kwh * (1 - dane.autokonsumpcja_pv_procent / 100)
    nadwyzka_pv_dzienna = nadwyzka_pv_roczna / 365

    # Peak shaving: zakładamy 3h szczytu
    peak_shaving_kwh = dane.moc_zamowiona_kw * 0.3 * 3  # 30% redukcji szczytu × 3h

    # Pojemność = max z nadwyżki PV i peak shaving, z korektą na degradację
    pojemnosc_bazowa = max(nadwyzka_pv_dzienna, peak_shaving_kwh)

    # Korekta na degradację (125% pojemności początkowej)
    pojemnosc_z_korekcja = pojemnosc_bazowa * 1.25

    # Zaokrąglij do najbliższych 50 kWh w górę
    pojemnosc = max(50, round(pojemnosc_z_korekcja / 50 + 0.5) * 50)

    # Moc = pojemność / 2 (stosunek C/2 — 2h pełnego ładowania/rozładowania)
    moc = pojemnosc / 2

    return ParametryBESS(
        pojemnosc_kwh=pojemnosc,
        moc_kw=moc,
        rte=0.90,
        cykle_dzienne=1.0,
        koszt_pln_kwh=2000,  # średnia cena C&I w Polsce 2026 (spadek 31% r/r)
        zywotnosc_lat=15,
        degradacja_roczna=2.0,
    )


def oblicz_oszczednosci(dane: DaneFaktury, bess: ParametryBESS) -> WynikiKalkulacji:
    """Oblicza oszczędności z instalacji BESS."""

    # === CAPEX ===
    capex_bess = bess.pojemnosc_kwh * bess.koszt_pln_kwh
    capex_ems = 30_000  # EMS software + konfiguracja
    capex_instalacja = capex_bess * 0.10  # ~10% na instalację, okablowanie, zabezpieczenia
    capex_calkowity = capex_bess + capex_ems + capex_instalacja

    # === 1. AUTOKONSUMPCJA PV ===
    # Nadwyżki PV kierowane do magazynu zamiast do sieci
    nadwyzka_pv_roczna = dane.roczna_produkcja_pv_kwh * (1 - dane.autokonsumpcja_pv_procent / 100)
    pojemnosc_roczna_magazynu = bess.pojemnosc_kwh * 365 * bess.rte  # ile magazyn może przetworzyć rocznie

    # Ile nadwyżek PV trafia do magazynu (min z nadwyżki i pojemności)
    nadwyzka_do_magazynu = min(nadwyzka_pv_roczna, pojemnosc_roczna_magazynu * 0.6)  # 60% pojemności na PV

    # Wartość: różnica między ceną autokonsumpcji a ceną net-billing
    cena_autokonsumpcji = dane.cena_energii_pln_kwh + dane.oplata_dystrybucyjna_pln_kwh  # pełna wartość
    cena_net_billing = dane.cena_energii_pln_kwh * 0.6  # ~60% ceny (RCE w godzinach PV × 1,23)

    roznica_na_kwh = cena_autokonsumpcji - cena_net_billing
    oszczednosc_autokonsumpcja = nadwyzka_do_magazynu * roznica_na_kwh

    # Nowa autokonsumpcja
    nowa_autokonsumpcja_kwh = (dane.roczna_produkcja_pv_kwh * dane.autokonsumpcja_pv_procent / 100
                               + nadwyzka_do_magazynu)
    nowa_autokonsumpcja_procent = min(95, (nowa_autokonsumpcja_kwh / dane.roczna_produkcja_pv_kwh) * 100)

    # === 2. ARBITRAŻ CENOWY ===
    # Spread: kupno po niskich cenach TGE, unikanie zakupu w szczycie
    # Średni spread umiarkowany: 0.30 PLN/kWh (300 PLN/MWh) – typowy spread peak vs off-peak 2025-2026
    # W dni z ujemnymi cenami spread może sięgać 0.60+ PLN/kWh
    spread_sredni = 0.30  # PLN/kWh
    # Magazyn wykonuje ~0.7 cyklu dziennie na arbitrażu (pozostałe 0.3 na autokonsumpcję PV)
    # Ale część pojemności jest już użyta na PV, więc liczymy osobno
    energia_arbitraz_dzienna = bess.pojemnosc_kwh * bess.rte * 0.5  # 50% pojemności na arbitraż
    # Nie codziennie opłaca się arbitraż – ok. 300 dni w roku (weekendy, święta, słabe spready)
    oszczednosc_arbitraz = energia_arbitraz_dzienna * spread_sredni * 300

    # === 3. PEAK SHAVING (opłata mocowa) ===
    # Obecna opłata mocowa roczna
    zuzycie_szczytu = dane.roczne_zuzycie_kwh * 0.65  # ~65% zużycia w godz. 7-22

    kategorie_mnozniki = {'K1': 0.17, 'K2': 0.40, 'K3': 0.70, 'K4': 1.00}
    mnoznik_obecny = kategorie_mnozniki.get(dane.kategoria_mocowa, 1.0)
    oplata_mocowa_obecna = zuzycie_szczytu / 1000 * dane.oplata_mocowa_pln_mwh * mnoznik_obecny

    # Po instalacji BESS: wygładzenie profilu → przejście o 1-2 kategorie niżej
    if dane.kategoria_mocowa == 'K4':
        nowa_kategoria = 'K2'
    elif dane.kategoria_mocowa == 'K3':
        nowa_kategoria = 'K1'
    elif dane.kategoria_mocowa == 'K2':
        nowa_kategoria = 'K1'
    else:
        nowa_kategoria = 'K1'

    mnoznik_nowy = kategorie_mnozniki[nowa_kategoria]
    oplata_mocowa_nowa = zuzycie_szczytu / 1000 * dane.oplata_mocowa_pln_mwh * mnoznik_nowy
    oszczednosc_peak_shaving = oplata_mocowa_obecna - oplata_mocowa_nowa

    # === SUMA ===
    oszczednosc_calkowita = oszczednosc_autokonsumpcja + oszczednosc_arbitraz + oszczednosc_peak_shaving

    # === ZWROT INWESTYCJI ===
    okres_zwrotu = capex_calkowity / oszczednosc_calkowita if oszczednosc_calkowita > 0 else 99

    # ROI 10 lat (uproszczony, bez dyskonta)
    przychod_10lat = oszczednosc_calkowita * 10
    roi_10lat = ((przychod_10lat - capex_calkowity) / capex_calkowity) * 100

    # NPV 10 lat (stopa dyskonta 6%)
    stopa = 0.06
    npv = -capex_calkowity
    for rok in range(1, 11):
        # Korekta na degradację
        degradacja = (1 - bess.degradacja_roczna / 100) ** rok
        npv += oszczednosc_calkowita * degradacja / (1 + stopa) ** rok

    # Redukcja rachunku
    roczny_rachunek = dane.sredni_rachunek_miesiac_pln * 12
    redukcja_rachunku = (oszczednosc_calkowita / roczny_rachunek * 100) if roczny_rachunek > 0 else 0

    # CO2 (współczynnik emisji PL: ~0.7 kg CO2/kWh z sieci)
    energia_zaoszczedzona = nadwyzka_do_magazynu + energia_arbitraz_dzienna * 365
    oszczednosc_co2 = energia_zaoszczedzona * 0.0007  # tony

    return WynikiKalkulacji(
        capex_bess_pln=capex_bess,
        capex_ems_pln=capex_ems,
        capex_instalacja_pln=capex_instalacja,
        capex_calkowity_pln=capex_calkowity,
        oszczednosc_autokonsumpcja_pln=oszczednosc_autokonsumpcja,
        oszczednosc_arbitraz_pln=oszczednosc_arbitraz,
        oszczednosc_peak_shaving_pln=oszczednosc_peak_shaving,
        oszczednosc_calkowita_roczna_pln=oszczednosc_calkowita,
        okres_zwrotu_lat=okres_zwrotu,
        roi_10lat_procent=roi_10lat,
        npv_10lat_pln=npv,
        nowa_autokonsumpcja_procent=nowa_autokonsumpcja_procent,
        redukcja_rachunku_procent=redukcja_rachunku,
        oszczednosc_co2_ton_rok=oszczednosc_co2,
    )


def generuj_xlsx(dane: DaneFaktury, bess: ParametryBESS, wyniki: WynikiKalkulacji, sciezka: str):
    """Generuje plik XLSX z ofertą."""

    wb = Workbook()

    # === Style ===
    header_font = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
    header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    subheader_font = Font(name='Calibri', bold=True, size=11, color='003366')
    subheader_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    value_font = Font(name='Calibri', size=11)
    money_font = Font(name='Calibri', size=11, bold=True, color='006600')
    warning_font = Font(name='Calibri', size=11, bold=True, color='CC0000')
    highlight_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    def set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def add_header_row(ws, row, text, cols=3):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 35

    def add_subheader(ws, row, text, cols=3):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = subheader_font
        cell.fill = subheader_fill

    def add_row(ws, row, label, value, unit='', bold_value=False, highlight=False):
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = value_font
        c1.border = thin_border

        c2 = ws.cell(row=row, column=2, value=value)
        c2.font = money_font if bold_value else value_font
        c2.border = thin_border
        c2.alignment = Alignment(horizontal='right')
        if isinstance(value, float):
            c2.number_format = '#,##0.00' if value < 100 else '#,##0'

        c3 = ws.cell(row=row, column=3, value=unit)
        c3.font = value_font
        c3.border = thin_border

        if highlight:
            for c in [c1, c2, c3]:
                c.fill = highlight_fill

    # ===== ARKUSZ 1: OFERTA =====
    ws1 = wb.active
    ws1.title = 'Oferta BESS'
    set_col_widths(ws1, [45, 20, 15])

    row = 1
    add_header_row(ws1, row, f'OFERTA: Magazyn Energii BESS dla {dane.nazwa_firmy}')
    row += 2

    # Dane klienta
    add_subheader(ws1, row, 'DANE KLIENTA')
    row += 1
    client_data = [
        ('Nazwa firmy', dane.nazwa_firmy, ''),
        ('Roczne zużycie energii', dane.roczne_zuzycie_kwh, 'kWh/rok'),
        ('Moc zamówiona', dane.moc_zamowiona_kw, 'kW'),
        ('Moc instalacji PV', dane.moc_pv_kwp, 'kWp'),
        ('Roczna produkcja PV', dane.roczna_produkcja_pv_kwh, 'kWh/rok'),
        ('Obecna autokonsumpcja PV', dane.autokonsumpcja_pv_procent, '%'),
        ('Średnia cena energii', dane.cena_energii_pln_kwh, 'PLN/kWh'),
        ('Średni rachunek miesięczny', dane.sredni_rachunek_miesiac_pln, 'PLN/mies.'),
        ('Kategoria opłaty mocowej', dane.kategoria_mocowa, ''),
    ]
    for label, val, unit in client_data:
        add_row(ws1, row, label, val, unit)
        row += 1

    row += 1
    add_subheader(ws1, row, 'PROPONOWANY SYSTEM BESS')
    row += 1
    bess_data = [
        ('Pojemność magazynu', bess.pojemnosc_kwh, 'kWh'),
        ('Moc magazynu', bess.moc_kw, 'kW'),
        ('Technologia', 'LFP (LiFePO₄)', ''),
        ('Sprawność (Round-Trip Efficiency)', bess.rte * 100, '%'),
        ('Żywotność', bess.zywotnosc_lat, 'lat'),
        ('Integracja', 'AC-coupled z istniejącym PV', ''),
        ('System EMS', 'Z optymalizacją cen TGE (15 min)', ''),
        ('Taryfa', 'Dynamiczna (ceny 15-minutowe TGE)', ''),
    ]
    for label, val, unit in bess_data:
        add_row(ws1, row, label, val, unit)
        row += 1

    row += 1
    add_subheader(ws1, row, 'KOSZT INWESTYCJI (CAPEX)')
    row += 1
    capex_data = [
        ('Magazyn energii BESS', wyniki.capex_bess_pln, 'PLN netto'),
        ('Oprogramowanie EMS', wyniki.capex_ems_pln, 'PLN netto'),
        ('Instalacja, okablowanie, zabezpieczenia', wyniki.capex_instalacja_pln, 'PLN netto'),
    ]
    for label, val, unit in capex_data:
        add_row(ws1, row, label, val, unit)
        row += 1
    add_row(ws1, row, 'CAPEX CAŁKOWITY', wyniki.capex_calkowity_pln, 'PLN netto', bold_value=True, highlight=True)
    row += 2

    add_subheader(ws1, row, 'OSZCZĘDNOŚCI ROCZNE')
    row += 1
    savings_data = [
        ('Autokonsumpcja PV (nadwyżki → magazyn)', wyniki.oszczednosc_autokonsumpcja_pln, 'PLN/rok'),
        ('Arbitraż cenowy (taryfa dynamiczna TGE)', wyniki.oszczednosc_arbitraz_pln, 'PLN/rok'),
        ('Peak shaving (redukcja opłaty mocowej)', wyniki.oszczednosc_peak_shaving_pln, 'PLN/rok'),
    ]
    for label, val, unit in savings_data:
        add_row(ws1, row, label, val, unit, bold_value=True)
        row += 1
    add_row(ws1, row, 'OSZCZĘDNOŚĆ ROCZNA ŁĄCZNIE', wyniki.oszczednosc_calkowita_roczna_pln, 'PLN/rok', bold_value=True, highlight=True)
    row += 2

    add_subheader(ws1, row, 'ZWROT Z INWESTYCJI')
    row += 1
    roi_data = [
        ('Okres zwrotu (prosty)', round(wyniki.okres_zwrotu_lat, 1), 'lat'),
        ('ROI w 10 lat', round(wyniki.roi_10lat_procent, 0), '%'),
        ('NPV w 10 lat (dyskonto 6%)', round(wyniki.npv_10lat_pln, 0), 'PLN'),
        ('Nowa autokonsumpcja PV', round(wyniki.nowa_autokonsumpcja_procent, 0), '%'),
        ('Redukcja rachunku za energię', round(wyniki.redukcja_rachunku_procent, 0), '%'),
        ('Redukcja emisji CO₂', round(wyniki.oszczednosc_co2_ton_rok, 1), 'ton/rok'),
    ]
    for label, val, unit in roi_data:
        add_row(ws1, row, label, val, unit, bold_value=True, highlight=True)
        row += 1

    row += 2
    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    note_cell = ws1.cell(row=row, column=1,
                         value='* Kalkulacja ma charakter szacunkowy. Rzeczywiste oszczędności zależą od profilu zużycia, '
                               'warunków pogodowych i cen rynkowych. Zalecamy szczegółową symulację z danymi 15-minutowymi.')
    note_cell.font = Font(name='Calibri', size=9, italic=True, color='666666')
    note_cell.alignment = Alignment(wrap_text=True)
    ws1.row_dimensions[row].height = 40

    # ===== ARKUSZ 2: ANALIZA 10-LETNIA =====
    ws2 = wb.create_sheet('Analiza 10-letnia')
    set_col_widths(ws2, [8, 20, 20, 20, 20, 20])

    row = 1
    add_header_row(ws2, row, 'ANALIZA FINANSOWA – PROJEKCJA 10-LETNIA', cols=6)
    row += 2

    # Nagłówki tabeli
    headers = ['Rok', 'Oszczędność\nroczna (PLN)', 'Degradacja\nmagazynu (%)', 'Oszczędność\npo degradacji',
               'Cash flow\nskumulowany', 'Wartość\nmagazynu (kWh)']
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=row, column=col, value=h)
        cell.font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
        cell.fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws2.row_dimensions[row].height = 40
    row += 1

    # Rok 0
    for col, val in enumerate([0, 0, 100, -wyniki.capex_calkowity_pln, -wyniki.capex_calkowity_pln, bess.pojemnosc_kwh], 1):
        cell = ws2.cell(row=row, column=col, value=val)
        cell.font = value_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='right')
        if isinstance(val, float) and abs(val) > 100:
            cell.number_format = '#,##0'
    row += 1

    # Lata 1-10
    skumulowany = -wyniki.capex_calkowity_pln
    for rok in range(1, 11):
        degradacja = (1 - bess.degradacja_roczna / 100) ** rok
        oszcz_po_degradacji = wyniki.oszczednosc_calkowita_roczna_pln * degradacja
        skumulowany += oszcz_po_degradacji
        pojemnosc_aktualna = bess.pojemnosc_kwh * degradacja

        vals = [rok, wyniki.oszczednosc_calkowita_roczna_pln, round(degradacja * 100, 1),
                round(oszcz_po_degradacji, 0), round(skumulowany, 0), round(pojemnosc_aktualna, 0)]

        for col, val in enumerate(vals, 1):
            cell = ws2.cell(row=row, column=col, value=val)
            cell.font = value_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='right')
            if isinstance(val, float) and abs(val) > 100:
                cell.number_format = '#,##0'

            # Podświetl rok zwrotu
            if col == 5 and skumulowany >= 0:
                cell.fill = highlight_fill
                cell.font = money_font

        row += 1

    # ===== ARKUSZ 3: KORZYŚCI =====
    ws3 = wb.create_sheet('Korzyści')
    set_col_widths(ws3, [50, 25])

    row = 1
    add_header_row(ws3, row, 'KORZYŚCI Z INSTALACJI MAGAZYNU ENERGII', cols=2)
    row += 2

    korzysci = [
        ('KORZYŚCI FINANSOWE', [
            f'Roczna oszczędność: {wyniki.oszczednosc_calkowita_roczna_pln:,.0f} PLN',
            f'Zwrot inwestycji w: {wyniki.okres_zwrotu_lat:.1f} lat',
            f'Redukcja rachunku za energię o: {wyniki.redukcja_rachunku_procent:.0f}%',
            f'ROI w 10 lat: {wyniki.roi_10lat_procent:.0f}%',
        ]),
        ('KORZYŚCI OPERACYJNE', [
            'Zakup energii po cenach 15-minutowych z TGE (taryfa dynamiczna)',
            'Automatyczne ładowanie w godzinach ujemnych/niskich cen (10:00-15:00)',
            'Rozładowanie w szczycie cenowym (17:00-21:00) – unikanie drogiej energii',
            'Wzrost niezależności energetycznej – ochrona przed blackoutami',
        ]),
        ('KORZYŚCI Z PV', [
            f'Wzrost autokonsumpcji PV z {dane.autokonsumpcja_pv_procent:.0f}% do {wyniki.nowa_autokonsumpcja_procent:.0f}%',
            'Nadwyżki PV do magazynu zamiast do sieci po niskich cenach net-billing',
            'Pełne wykorzystanie istniejącej instalacji PV',
            'Wartość autokonsumowanej energii 2-3× wyższa niż net-billing',
        ]),
        ('KORZYŚCI ŚRODOWISKOWE', [
            f'Redukcja emisji CO₂: {wyniki.oszczednosc_co2_ton_rok:.1f} ton/rok',
            'Wsparcie transformacji energetycznej zakładu',
            'Lepszy wizerunek firmy (ESG, zielona energia)',
            'Przygotowanie na przyszłe regulacje klimatyczne',
        ]),
        ('KORZYŚCI RYNKOWE (przyszłość)', [
            'Możliwość świadczenia usług pomocniczych (aFRR, mFRR) – dodatkowe przychody',
            'Peak shaving – redukcja opłaty mocowej (wzrost o 55% w 2026)',
            'Hedge przed rosnącymi cenami energii',
            'Uczestnictwo w rynku bilansującym',
        ]),
    ]

    for section_title, items in korzysci:
        ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        cell = ws3.cell(row=row, column=1, value=section_title)
        cell.font = subheader_font
        cell.fill = subheader_fill
        row += 1
        for item in items:
            ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            cell = ws3.cell(row=row, column=1, value=f'  ✓  {item}')
            cell.font = value_font
            row += 1
        row += 1

    # Zapisz
    wb.save(sciezka)
    print(f'\nOferta zapisana: {sciezka}')


def pobierz_dane_interaktywnie() -> DaneFaktury:
    """Pobiera dane z faktury interaktywnie od użytkownika."""

    print('\n' + '=' * 60)
    print('  KALKULATOR BESS – Magazyn Energii dla Zakładów z PV')
    print('=' * 60)
    print('\nProszę podać dane z faktury za energię elektryczną.\n')

    def ask_float(prompt, default=None):
        while True:
            try:
                suffix = f' [{default}]' if default is not None else ''
                val = input(f'  {prompt}{suffix}: ').strip()
                if not val and default is not None:
                    return float(default)
                return float(val.replace(',', '.').replace(' ', ''))
            except ValueError:
                print('  ⚠ Podaj liczbę (np. 150000)')

    def ask_str(prompt, default=None):
        suffix = f' [{default}]' if default else ''
        val = input(f'  {prompt}{suffix}: ').strip()
        return val if val else (default or '')

    nazwa = ask_str('Nazwa firmy', 'Przykładowa Sp. z o.o.')
    print()

    print('--- Dane zużycia ---')
    zuzycie = ask_float('Roczne zużycie energii (kWh/rok)')
    moc_zam = ask_float('Moc zamówiona (kW)')
    rachunek = ask_float('Średni rachunek miesięczny (PLN/mies.)')
    print()

    print('--- Dane instalacji PV ---')
    moc_pv = ask_float('Moc instalacji PV (kWp)')
    produkcja_pv = ask_float(f'Roczna produkcja PV (kWh/rok)', round(moc_pv * 1050))
    autokons = ask_float('Obecna autokonsumpcja PV (%)', 35)
    print()

    print('--- Ceny energii ---')
    cena = ask_float('Średnia cena energii (PLN/kWh)', 0.65)
    dystr = ask_float('Opłata dystrybucyjna (PLN/kWh)', 0.25)
    mocowa = ask_float('Stawka opłaty mocowej (PLN/MWh)', 219.40)
    print()

    print('  Kategoria opłaty mocowej:')
    print('    K1 = profil płaski (<5% różnicy), K2 = 5-10%, K3 = 10-15%, K4 = >15%')
    kat = ask_str('Kategoria (K1/K2/K3/K4)', 'K3').upper()
    if kat not in ['K1', 'K2', 'K3', 'K4']:
        kat = 'K3'

    return DaneFaktury(
        nazwa_firmy=nazwa,
        roczne_zuzycie_kwh=zuzycie,
        moc_zamowiona_kw=moc_zam,
        moc_pv_kwp=moc_pv,
        roczna_produkcja_pv_kwh=produkcja_pv,
        autokonsumpcja_pv_procent=autokons,
        cena_energii_pln_kwh=cena,
        oplata_dystrybucyjna_pln_kwh=dystr,
        oplata_mocowa_pln_mwh=mocowa,
        kategoria_mocowa=kat,
        sredni_rachunek_miesiac_pln=rachunek,
    )


def dane_demo() -> DaneFaktury:
    """Dane przykładowe dla trybu demo."""
    return DaneFaktury(
        nazwa_firmy='Przykładowy Zakład Produkcyjny Sp. z o.o.',
        roczne_zuzycie_kwh=500_000,
        moc_zamowiona_kw=200,
        moc_pv_kwp=150,
        roczna_produkcja_pv_kwh=157_500,
        autokonsumpcja_pv_procent=35,
        cena_energii_pln_kwh=0.65,
        oplata_dystrybucyjna_pln_kwh=0.25,
        oplata_mocowa_pln_mwh=219.40,
        kategoria_mocowa='K3',
        sredni_rachunek_miesiac_pln=35_000,
    )


def main():
    demo = '--demo' in sys.argv

    if demo:
        print('\n🔋 Tryb DEMO – używam danych przykładowych\n')
        dane = dane_demo()
    else:
        dane = pobierz_dane_interaktywnie()

    # Oblicz parametry BESS
    bess = oblicz_parametry_bess(dane)

    # Oblicz oszczędności
    wyniki = oblicz_oszczednosci(dane, bess)

    # Wyświetl podsumowanie
    print('\n' + '=' * 60)
    print('  WYNIKI KALKULACJI')
    print('=' * 60)
    print(f'\n  Firma: {dane.nazwa_firmy}')
    print(f'\n  --- Proponowany magazyn ---')
    print(f'  Pojemność:  {bess.pojemnosc_kwh:.0f} kWh')
    print(f'  Moc:        {bess.moc_kw:.0f} kW')
    print(f'  Technologia: LFP (AC-coupled)')
    print(f'\n  --- CAPEX ---')
    print(f'  BESS:        {wyniki.capex_bess_pln:>12,.0f} PLN')
    print(f'  EMS:         {wyniki.capex_ems_pln:>12,.0f} PLN')
    print(f'  Instalacja:  {wyniki.capex_instalacja_pln:>12,.0f} PLN')
    print(f'  RAZEM:       {wyniki.capex_calkowity_pln:>12,.0f} PLN')
    print(f'\n  --- Oszczędności roczne ---')
    print(f'  Autokonsumpcja PV: {wyniki.oszczednosc_autokonsumpcja_pln:>10,.0f} PLN/rok')
    print(f'  Arbitraż cenowy:   {wyniki.oszczednosc_arbitraz_pln:>10,.0f} PLN/rok')
    print(f'  Peak shaving:      {wyniki.oszczednosc_peak_shaving_pln:>10,.0f} PLN/rok')
    print(f'  RAZEM:             {wyniki.oszczednosc_calkowita_roczna_pln:>10,.0f} PLN/rok')
    print(f'\n  --- Zwrot ---')
    print(f'  Okres zwrotu:       {wyniki.okres_zwrotu_lat:.1f} lat')
    print(f'  ROI (10 lat):       {wyniki.roi_10lat_procent:.0f}%')
    print(f'  NPV (10 lat, 6%):   {wyniki.npv_10lat_pln:,.0f} PLN')
    print(f'  Autokonsumpcja PV:  {dane.autokonsumpcja_pv_procent:.0f}% → {wyniki.nowa_autokonsumpcja_procent:.0f}%')
    print(f'  Redukcja rachunku:  {wyniki.redukcja_rachunku_procent:.0f}%')

    # Generuj XLSX
    sciezka = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           f'Oferta_BESS_{dane.nazwa_firmy.replace(" ", "_")[:30]}.xlsx')
    generuj_xlsx(dane, bess, wyniki, sciezka)

    print(f'\n  ✅ Oferta wygenerowana!')
    print(f'  📄 Plik: {sciezka}')
    print()


if __name__ == '__main__':
    main()
