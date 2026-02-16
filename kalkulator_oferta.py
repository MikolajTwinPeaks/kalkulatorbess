#!/usr/bin/env python3
"""
Kompleksowy kalkulator ofertowy: Energia Elektryczna + Gaz + PV + BESS + DSR + KMB.

Generuje pełną ofertę XLSX z analizą:
- Kontraktu na ee (FIX/RDN/MIX)
- Kosztów dystrybucji
- Rekomendacji PV + sizing
- Rekomendacji BESS + arbitraż + peak shaving
- DSR (Demand Side Response)
- KMB (Kompensacja Mocy Biernej)
- Opcji finansowania (leasing, ESCO, PPA, kredyt, raty)
- ROI dla każdego produktu

Użycie:
    python3 kalkulator_oferta.py          # tryb interaktywny
    python3 kalkulator_oferta.py --demo   # dane przykładowe
"""

import sys
import os
import io
import math
import tempfile
from dataclasses import dataclass, field
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart


# ============================================================
# DATA MODELS
# ============================================================

@dataclass
class DaneKlienta:
    nazwa_firmy: str
    nip: str
    branza: str
    dni_pracy: str                      # np. "Pn-Pt"
    godziny_pracy: str                  # np. "6:00-22:00 (2 zmiany)"
    # Energia elektryczna
    roczne_zuzycie_ee_kwh: float
    moc_umowna_kw: float
    moc_przylaczeniowa_kw: float
    grupa_taryfowa: str                 # np. C22a, B21
    osd: str                            # np. Tauron, Enea
    sredni_rachunek_ee_mies_pln: float
    cena_ee_pln_kwh: float              # cena energii z obecnej umowy
    oplata_dystr_pln_kwh: float
    oplata_mocowa_pln_mwh: float
    kategoria_mocowa: str               # K1-K4
    data_konca_umowy_ee: str
    typ_umowy_ee: str                   # FIX / RDN / MIX
    # Gaz
    roczne_zuzycie_gaz_kwh: float
    sredni_rachunek_gaz_mies_pln: float
    cena_gaz_pln_kwh: float
    data_konca_umowy_gaz: str
    # PV istniejące
    ma_pv: bool
    moc_pv_kwp: float
    roczna_produkcja_pv_kwh: float
    autokonsumpcja_pv_procent: float
    # Inne
    ma_kmb: bool
    moc_bierna_kvar: float              # jeśli ma KMB
    ma_agregat: bool
    potrzebuje_go: bool                 # gwarancje pochodzenia
    powierzchnia_dachu_m2: float        # dostępna na PV
    # Moc bierna z FV
    wspolczynnik_cos_phi: float         # np. 0.85


@dataclass
class RekomendacjaEE:
    """Rekomendacja produktu na energię elektryczną."""
    produkt: str                        # FIX / RDN / MIX
    cena_fix_pln_kwh: float
    cena_rdn_srednia_pln_kwh: float
    cena_mix_pln_kwh: float             # np. 50% FIX + 50% RDN
    oszczednosc_roczna_pln: float
    uzasadnienie: str


@dataclass
class RekomendacjaPV:
    """Rekomendacja instalacji PV."""
    nowa_moc_kwp: float                 # dodatkowa moc PV
    roczna_produkcja_kwh: float
    capex_pln: float
    oszczednosc_roczna_pln: float
    okres_zwrotu_lat: float
    autokonsumpcja_procent: float


@dataclass
class RekomendacjaBESS:
    """Rekomendacja magazynu energii."""
    pojemnosc_kwh: float
    moc_kw: float
    capex_pln: float
    oszczednosc_autokonsumpcja_pln: float
    oszczednosc_arbitraz_pln: float
    oszczednosc_peak_shaving_pln: float
    oszczednosc_calkowita_pln: float
    okres_zwrotu_lat: float


@dataclass
class RekomendacjaDSR:
    """Rekomendacja DSR (Demand Side Response)."""
    potencjal_redukcji_kw: float
    przychod_roczny_pln: float
    koszt_wdrozenia_pln: float
    uzasadnienie: str


@dataclass
class RekomendacjaKMB:
    """Rekomendacja KMB (Kompensacja Mocy Biernej)."""
    moc_potrzebna_kvar: float
    capex_pln: float
    oszczednosc_roczna_pln: float
    okres_zwrotu_lat: float


@dataclass
class OpcjaFinansowania:
    nazwa: str
    opis: str
    wklad_wlasny_procent: float
    rata_miesieczna_pln: float
    okres_mies: int
    koszt_calkowity_pln: float
    korzysc_podatkowa_roczna_pln: float
    uwagi: str


# ============================================================
# CALCULATION ENGINE
# ============================================================

def oblicz_rekomendacje_ee(dane: DaneKlienta) -> RekomendacjaEE:
    """Rekomendacja produktu energetycznego FIX/RDN/MIX."""

    # Ceny rynkowe 2026
    cena_fix = 0.58         # PLN/kWh – cena FIX rynkowa 2026
    cena_rdn_srednia = 0.50 # PLN/kWh – średnia RDN (z potencjałem optymalizacji)
    cena_mix = 0.54         # 50% FIX + 50% RDN

    obecna_cena = dane.cena_ee_pln_kwh
    zuzycie = dane.roczne_zuzycie_ee_kwh

    oszcz_fix = max(0, (obecna_cena - cena_fix) * zuzycie)
    oszcz_rdn = max(0, (obecna_cena - cena_rdn_srednia) * zuzycie)
    oszcz_mix = max(0, (obecna_cena - cena_mix) * zuzycie)

    # Rekomendacja na podstawie profilu
    godziny = dane.godziny_pracy.lower()
    if '24' in godziny or '3 zmian' in godziny:
        # Praca ciągła – RDN opłacalny (można wykorzystać tanie godziny nocne)
        produkt = 'RDN'
        oszczednosc = oszcz_rdn
        uzasadnienie = ('Zakład pracuje w trybie ciągłym (24h/3 zmiany) – duży potencjał przesuwania '
                        'zużycia na tanie godziny. Taryfa dynamiczna RDN z cenami 15-minutowymi TGE '
                        'pozwala kupować energię w godzinach niskich cen (10:00-15:00, noc).')
    elif dane.ma_pv or dane.moc_pv_kwp > 0:
        # Ma PV – MIX lub RDN (PV pokrywa dzień, RDN na resztę)
        produkt = 'MIX'
        oszczednosc = oszcz_mix
        uzasadnienie = ('Zakład posiada instalację PV – MIX (50% FIX + 50% RDN) daje stabilność '
                        'cenową z możliwością korzystania z niskich cen w godzinach generacji PV. '
                        'Z magazynem BESS rekomendacja zmienia się na RDN.')
    else:
        # Standardowy profil – FIX dla bezpieczeństwa
        if obecna_cena > cena_fix * 1.1:
            produkt = 'FIX'
            oszczednosc = oszcz_fix
            uzasadnienie = ('Obecna cena jest wyższa od rynkowej ceny FIX. Rekomendujemy FIX '
                            'dla przewidywalności kosztów. Po instalacji PV+BESS warto rozważyć MIX lub RDN.')
        else:
            produkt = 'MIX'
            oszczednosc = oszcz_mix
            uzasadnienie = ('MIX łączy stabilność FIX z potencjałem oszczędności RDN. '
                            'Rekomendowany dla zakładów z profilem 1-2 zmianowym.')

    return RekomendacjaEE(
        produkt=produkt,
        cena_fix_pln_kwh=cena_fix,
        cena_rdn_srednia_pln_kwh=cena_rdn_srednia,
        cena_mix_pln_kwh=cena_mix,
        oszczednosc_roczna_pln=oszczednosc,
        uzasadnienie=uzasadnienie,
    )


def oblicz_rekomendacje_pv(dane: DaneKlienta) -> Optional[RekomendacjaPV]:
    """Rekomendacja dodatkowej/nowej instalacji PV."""

    if dane.powierzchnia_dachu_m2 < 20:
        return None

    # 1 kWp wymaga ok. 5-6 m² dachu
    max_moc_z_dachu = dane.powierzchnia_dachu_m2 / 5.5
    # Optymalna moc: pokrycie ~70% rocznego zużycia ee (z uwzględnieniem istniejącej PV)
    istniejaca_produkcja = dane.roczna_produkcja_pv_kwh if dane.ma_pv else 0
    potrzebna_produkcja = dane.roczne_zuzycie_ee_kwh * 0.70 - istniejaca_produkcja

    if potrzebna_produkcja <= 0:
        return None

    # 1 kWp produkuje ok. 1000-1100 kWh/rok w Polsce
    potrzebna_moc = potrzebna_produkcja / 1050
    nowa_moc = min(potrzebna_moc, max_moc_z_dachu)
    nowa_moc = max(10, round(nowa_moc / 5) * 5)  # zaokrąglij do 5 kWp

    roczna_produkcja = nowa_moc * 1050
    capex_pln_kwp = 3800 if nowa_moc < 50 else 3200 if nowa_moc < 200 else 2800
    capex = nowa_moc * capex_pln_kwp

    # Autokonsumpcja (bez magazynu): 30-50% zależnie od profilu
    if '24' in dane.godziny_pracy or '3 zmian' in dane.godziny_pracy.lower():
        autokons = 0.50
    elif 'sob' in dane.dni_pracy.lower() or '6' in dane.dni_pracy:
        autokons = 0.40
    else:
        autokons = 0.35

    # Oszczędność = autokonsumpcja × cena pełna + reszta × cena net-billing
    cena_pelna = dane.cena_ee_pln_kwh + dane.oplata_dystr_pln_kwh
    cena_net_billing = dane.cena_ee_pln_kwh * 0.5
    oszczednosc = roczna_produkcja * (autokons * cena_pelna + (1 - autokons) * cena_net_billing)

    okres_zwrotu = capex / oszczednosc if oszczednosc > 0 else 99

    return RekomendacjaPV(
        nowa_moc_kwp=nowa_moc,
        roczna_produkcja_kwh=roczna_produkcja,
        capex_pln=capex,
        oszczednosc_roczna_pln=oszczednosc,
        okres_zwrotu_lat=okres_zwrotu,
        autokonsumpcja_procent=autokons * 100,
    )


def oblicz_rekomendacje_bess(dane: DaneKlienta) -> RekomendacjaBESS:
    """Rekomendacja magazynu energii BESS."""

    # Łączna moc PV (istniejąca + rekomendowana)
    moc_pv_total = dane.moc_pv_kwp if dane.ma_pv else 0

    # Nadwyżka PV dziennie
    produkcja_pv = dane.roczna_produkcja_pv_kwh if dane.ma_pv else 0
    autokons = dane.autokonsumpcja_pv_procent / 100 if dane.ma_pv else 0
    nadwyzka_roczna = produkcja_pv * (1 - autokons)
    nadwyzka_dzienna = nadwyzka_roczna / 365

    # Peak shaving target: 30% redukcji mocy szczytowej × 3h
    peak_target = dane.moc_umowna_kw * 0.30 * 3

    # Arbitraż: wystarczająco na 1 pełny cykl
    arbitraz_target = dane.roczne_zuzycie_ee_kwh / 365 * 0.2  # 20% dziennego zużycia

    # Pojemność = max z trzech celów
    pojemnosc_bazowa = max(nadwyzka_dzienna, peak_target, arbitraz_target, 50)
    pojemnosc = max(50, round(pojemnosc_bazowa * 1.25 / 50 + 0.5) * 50)  # +25% degradacja
    moc = pojemnosc / 2

    # CAPEX
    koszt_kwh = 2000  # PLN/kWh (2026)
    capex_bess = pojemnosc * koszt_kwh
    capex_ems = 30000
    capex_inst = capex_bess * 0.10
    capex_total = capex_bess + capex_ems + capex_inst

    rte = 0.90

    # 1. Autokonsumpcja PV
    nadwyzka_do_magazynu = min(nadwyzka_roczna, pojemnosc * 365 * rte * 0.5)
    cena_autokons = dane.cena_ee_pln_kwh + dane.oplata_dystr_pln_kwh
    cena_net_bill = dane.cena_ee_pln_kwh * 0.5
    oszcz_autokons = nadwyzka_do_magazynu * (cena_autokons - cena_net_bill)

    # 2. Arbitraż cenowy
    spread = 0.30  # PLN/kWh
    energia_arbitraz = pojemnosc * rte * 0.5
    oszcz_arbitraz = energia_arbitraz * spread * 300  # 300 dni efektywnych

    # 3. Peak shaving
    zuzycie_szczytu = dane.roczne_zuzycie_ee_kwh * 0.65
    kat_mn = {'K1': 0.17, 'K2': 0.40, 'K3': 0.70, 'K4': 1.00}
    mn_obecny = kat_mn.get(dane.kategoria_mocowa, 1.0)
    oplata_obecna = zuzycie_szczytu / 1000 * dane.oplata_mocowa_pln_mwh * mn_obecny

    nowa_kat = {'K4': 'K2', 'K3': 'K1', 'K2': 'K1', 'K1': 'K1'}
    mn_nowy = kat_mn[nowa_kat[dane.kategoria_mocowa]]
    oplata_nowa = zuzycie_szczytu / 1000 * dane.oplata_mocowa_pln_mwh * mn_nowy
    oszcz_peak = oplata_obecna - oplata_nowa

    oszcz_total = oszcz_autokons + oszcz_arbitraz + oszcz_peak
    okres_zwrotu = capex_total / oszcz_total if oszcz_total > 0 else 99

    return RekomendacjaBESS(
        pojemnosc_kwh=pojemnosc,
        moc_kw=moc,
        capex_pln=capex_total,
        oszczednosc_autokonsumpcja_pln=oszcz_autokons,
        oszczednosc_arbitraz_pln=oszcz_arbitraz,
        oszczednosc_peak_shaving_pln=oszcz_peak,
        oszczednosc_calkowita_pln=oszcz_total,
        okres_zwrotu_lat=okres_zwrotu,
    )


def oblicz_rekomendacje_dsr(dane: DaneKlienta) -> RekomendacjaDSR:
    """Rekomendacja DSR (Demand Side Response)."""

    # DSR opłacalny od ~200 kW redukcji
    potencjal = dane.moc_umowna_kw * 0.15  # 15% mocy umownej
    potencjal = max(0, round(potencjal / 10) * 10)

    if potencjal < 50:
        return RekomendacjaDSR(
            potencjal_redukcji_kw=0,
            przychod_roczny_pln=0,
            koszt_wdrozenia_pln=0,
            uzasadnienie='Potencjał DSR poniżej 50 kW – nieopłacalny dla tego zakładu.',
        )

    # Przychód z DSR: ok. 200-400 PLN/kW/rok (rynek mocy + usługi pomocnicze)
    przychod_kw_rok = 300  # PLN/kW/rok (konserwatywnie)
    przychod_roczny = potencjal * przychod_kw_rok
    koszt_wdrozenia = 15000 + potencjal * 50  # system sterowania + integracja

    return RekomendacjaDSR(
        potencjal_redukcji_kw=potencjal,
        przychod_roczny_pln=przychod_roczny,
        koszt_wdrozenia_pln=koszt_wdrozenia,
        uzasadnienie=(
            f'Potencjał redukcji: {potencjal:.0f} kW (15% mocy umownej). '
            f'DSR pozwala na dodatkowe przychody z rynku mocy i usług pomocniczych PSE. '
            f'Wymagana jest możliwość redukcji obciążenia na wezwanie operatora (PSE) '
            f'lub agregatora (np. ALIANS). Z magazynem BESS potencjał DSR znacząco rośnie.'
        ),
    )


def oblicz_rekomendacje_kmb(dane: DaneKlienta) -> Optional[RekomendacjaKMB]:
    """Rekomendacja KMB (Kompensacja Mocy Biernej)."""

    if dane.ma_kmb:
        return None  # Już ma kompensację

    if dane.wspolczynnik_cos_phi >= 0.95:
        return None  # Nie potrzebuje

    # Oblicz potrzebną moc bierną
    moc_czynna = dane.moc_umowna_kw
    phi_obecny = math.acos(dane.wspolczynnik_cos_phi)
    phi_docelowy = math.acos(0.95)
    q_potrzebna = moc_czynna * (math.tan(phi_obecny) - math.tan(phi_docelowy))
    q_potrzebna = max(10, round(q_potrzebna / 5) * 5)

    # CAPEX: ok. 80-150 PLN/kvar dla kompensacji automatycznej
    capex = q_potrzebna * 120  # PLN

    # Oszczędność: opłata za moc bierną = ok. 5-15% rachunku dystrybucyjnego
    roczny_koszt_dystr = dane.roczne_zuzycie_ee_kwh * dane.oplata_dystr_pln_kwh
    oszczednosc = roczny_koszt_dystr * 0.10  # ~10% oszczędności na dystrybucji

    okres_zwrotu = capex / oszczednosc if oszczednosc > 0 else 99

    return RekomendacjaKMB(
        moc_potrzebna_kvar=q_potrzebna,
        capex_pln=capex,
        oszczednosc_roczna_pln=oszczednosc,
        okres_zwrotu_lat=okres_zwrotu,
    )


def oblicz_opcje_finansowania(capex_total: float) -> list[OpcjaFinansowania]:
    """Oblicza opcje finansowania dla łącznego CAPEX."""

    opcje = []

    # 1. Zakup za gotówkę
    amortyzacja_roczna = capex_total * 0.10  # 10% rocznie
    korzysc_cit = amortyzacja_roczna * 0.19  # CIT 19%
    opcje.append(OpcjaFinansowania(
        nazwa='Zakup za gotówkę',
        opis='Pełna płatność, własność od dnia 1. Amortyzacja 10%/rok (degresywna do 20%).',
        wklad_wlasny_procent=100,
        rata_miesieczna_pln=0,
        okres_mies=0,
        koszt_calkowity_pln=capex_total,
        korzysc_podatkowa_roczna_pln=korzysc_cit,
        uwagi='Najszybszy ROI. Pełna amortyzacja degresywna (20% w 1. roku).',
    ))

    # 2. Leasing operacyjny (84 mies.)
    okres_leasing = 84
    oprocentowanie_leasing = 0.065  # RRSO ~6.5%
    rata_leasing = capex_total * (oprocentowanie_leasing / 12 * (1 + oprocentowanie_leasing / 12) ** okres_leasing) / ((1 + oprocentowanie_leasing / 12) ** okres_leasing - 1)
    koszt_leasing = rata_leasing * okres_leasing + capex_total * 0.01  # + wykup 1%
    korzysc_leasing_op = rata_leasing * 12 * 0.19  # cała rata w KUP × CIT
    opcje.append(OpcjaFinansowania(
        nazwa='Leasing operacyjny (7 lat)',
        opis=f'Rata: {rata_leasing:,.0f} PLN/mies. netto. Cała rata = KUP. Wykup 1%.',
        wklad_wlasny_procent=0,
        rata_miesieczna_pln=rata_leasing,
        okres_mies=okres_leasing,
        koszt_calkowity_pln=koszt_leasing,
        korzysc_podatkowa_roczna_pln=korzysc_leasing_op,
        uwagi='0% wkładu własnego. Nie obciąża bilansu. mLeasing, PKO Leasing, EFL, BOS Leasing.',
    ))

    # 3. Leasing finansowy (120 mies.)
    okres_fin = 120
    rata_fin = capex_total * (oprocentowanie_leasing / 12 * (1 + oprocentowanie_leasing / 12) ** okres_fin) / ((1 + oprocentowanie_leasing / 12) ** okres_fin - 1)
    koszt_fin = rata_fin * okres_fin
    # Korzyść: amortyzacja + odsetki w KUP
    korzysc_fin = (amortyzacja_roczna + rata_fin * 12 * 0.35) * 0.19  # ~35% raty to odsetki na początku
    opcje.append(OpcjaFinansowania(
        nazwa='Leasing finansowy (10 lat)',
        opis=f'Rata: {rata_fin:,.0f} PLN/mies. netto. Amortyzacja + odsetki w KUP.',
        wklad_wlasny_procent=5,
        rata_miesieczna_pln=rata_fin,
        okres_mies=okres_fin,
        koszt_calkowity_pln=koszt_fin,
        korzysc_podatkowa_roczna_pln=korzysc_fin,
        uwagi='Środek trwały na bilansie. Można łączyć z dotacjami UE. Amortyzacja degresywna.',
    ))

    # 4. Kredyt ekologiczny BGK (FENG)
    premia = 0.50  # 50% kosztów kwalifikowanych (MŚP)
    kapital_kredytu = capex_total * (1 - premia)
    rata_bgk = kapital_kredytu * (0.07 / 12 * (1 + 0.07 / 12) ** 120) / ((1 + 0.07 / 12) ** 120 - 1)
    koszt_bgk = rata_bgk * 120
    opcje.append(OpcjaFinansowania(
        nazwa='Kredyt ekologiczny BGK (FENG)',
        opis=f'Premia ekologiczna: {premia*100:.0f}% ({capex_total*premia:,.0f} PLN). Kredyt na resztę.',
        wklad_wlasny_procent=0,
        rata_miesieczna_pln=rata_bgk,
        okres_mies=120,
        koszt_calkowity_pln=koszt_bgk,
        korzysc_podatkowa_roczna_pln=amortyzacja_roczna * 0.19,
        uwagi=f'Premia 50% dla MŚP (do 80% z bonusami). Wymagany audyt energetyczny i min. 30% redukcji zużycia energii. Nabory BGK.',
    ))

    # 5. ESCO (0 CAPEX)
    opcje.append(OpcjaFinansowania(
        nazwa='Model ESCO (EPC)',
        opis='ESCO finansuje inwestycję. Spłata z gwarantowanych oszczędności (10-15 lat).',
        wklad_wlasny_procent=0,
        rata_miesieczna_pln=0,
        okres_mies=180,  # 15 lat
        koszt_calkowity_pln=capex_total * 1.5,  # ESCO pobiera ~50% premii
        korzysc_podatkowa_roczna_pln=0,
        uwagi='Zero CAPEX. Ryzyko po stronie ESCO. Veolia ESCO Polska, ENGIE, Siemens. Dłuższy okres zwrotu.',
    ))

    # 6. PPA (on-site)
    opcje.append(OpcjaFinansowania(
        nazwa='On-site PPA (tylko PV)',
        opis='Developer buduje PV na terenie klienta. Klient kupuje energię po stałej cenie 10-15 lat.',
        wklad_wlasny_procent=0,
        rata_miesieczna_pln=0,
        okres_mies=180,
        koszt_calkowity_pln=0,
        korzysc_podatkowa_roczna_pln=0,
        uwagi='Cena PPA: 280-350 PLN/MWh (15-30% poniżej rynku). Dotyczy tylko PV. Polenergia, R.Power, Columbus.',
    ))

    return opcje


# ============================================================
# XLSX GENERATOR
# ============================================================

def generuj_oferte(dane: DaneKlienta, output_path: str):
    """Generuje pełną ofertę XLSX."""

    # Oblicz wszystkie rekomendacje
    rek_ee = oblicz_rekomendacje_ee(dane)
    rek_pv = oblicz_rekomendacje_pv(dane)
    rek_bess = oblicz_rekomendacje_bess(dane)
    rek_dsr = oblicz_rekomendacje_dsr(dane)
    rek_kmb = oblicz_rekomendacje_kmb(dane)

    # Łączny CAPEX
    capex_items = [('BESS', rek_bess.capex_pln)]
    if rek_pv:
        capex_items.append(('PV', rek_pv.capex_pln))
    if rek_dsr and rek_dsr.potencjal_redukcji_kw > 0:
        capex_items.append(('DSR', rek_dsr.koszt_wdrozenia_pln))
    if rek_kmb:
        capex_items.append(('KMB', rek_kmb.capex_pln))

    capex_total = sum(c for _, c in capex_items)
    opcje_fin = oblicz_opcje_finansowania(capex_total)

    # Łączna oszczędność roczna
    oszcz_items = [
        ('Kontrakt ee (zmiana taryfy)', rek_ee.oszczednosc_roczna_pln),
        ('BESS – autokonsumpcja PV', rek_bess.oszczednosc_autokonsumpcja_pln),
        ('BESS – arbitraż cenowy', rek_bess.oszczednosc_arbitraz_pln),
        ('BESS – peak shaving', rek_bess.oszczednosc_peak_shaving_pln),
    ]
    if rek_pv:
        oszcz_items.append(('PV – nowa instalacja', rek_pv.oszczednosc_roczna_pln))
    if rek_dsr and rek_dsr.przychod_roczny_pln > 0:
        oszcz_items.append(('DSR – przychód z redukcji', rek_dsr.przychod_roczny_pln))
    if rek_kmb:
        oszcz_items.append(('KMB – kompensacja', rek_kmb.oszczednosc_roczna_pln))

    oszcz_total = sum(o for _, o in oszcz_items)

    # ============ BUILD XLSX ============
    wb = Workbook()

    # Style
    hdr_font = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
    hdr_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    sec_font = Font(name='Calibri', bold=True, size=11, color='003366')
    sec_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    val_font = Font(name='Calibri', size=10)
    money_font = Font(name='Calibri', size=10, bold=True, color='006600')
    neg_font = Font(name='Calibri', size=10, bold=True, color='CC0000')
    hi_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    note_font = Font(name='Calibri', size=9, italic=True, color='666666')

    def sw(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def hdr(ws, r, txt, cols=4):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
        c = ws.cell(row=r, column=1, value=txt)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[r].height = 35

    def sec(ws, r, txt, cols=4):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
        c = ws.cell(row=r, column=1, value=txt)
        c.font = sec_font; c.fill = sec_fill

    def row2(ws, r, label, val, unit='', bold=False, hi=False):
        c1 = ws.cell(row=r, column=1, value=label); c1.font = val_font; c1.border = thin
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        c3 = ws.cell(row=r, column=3, value=val)
        c3.font = money_font if bold else val_font; c3.border = thin
        c3.alignment = Alignment(horizontal='right')
        if isinstance(val, (int, float)):
            c3.number_format = '#,##0' if abs(val) >= 10 else '#,##0.00'
        c4 = ws.cell(row=r, column=4, value=unit); c4.font = val_font; c4.border = thin
        if hi:
            for col in range(1, 5):
                ws.cell(row=r, column=col).fill = hi_fill

    # ====== ARK 1: PODSUMOWANIE OFERTY ======
    ws1 = wb.active
    ws1.title = 'Podsumowanie oferty'
    sw(ws1, [25, 20, 18, 15])

    r = 1
    hdr(ws1, r, f'OFERTA ENERGETYCZNA: {dane.nazwa_firmy}'); r += 2

    sec(ws1, r, 'DANE KLIENTA'); r += 1
    for label, val, unit in [
        ('Firma', dane.nazwa_firmy, ''),
        ('NIP', dane.nip, ''),
        ('Branża', dane.branza, ''),
        ('Profil pracy', f'{dane.dni_pracy}, {dane.godziny_pracy}', ''),
        ('Roczne zużycie ee', dane.roczne_zuzycie_ee_kwh, 'kWh/rok'),
        ('Roczne zużycie gazu', dane.roczne_zuzycie_gaz_kwh, 'kWh/rok'),
        ('Moc umowna', dane.moc_umowna_kw, 'kW'),
        ('Obecna cena ee', dane.cena_ee_pln_kwh, 'PLN/kWh'),
        ('Koniec umowy ee', dane.data_konca_umowy_ee, ''),
        ('Istniejące PV', f'{dane.moc_pv_kwp} kWp' if dane.ma_pv else 'BRAK', ''),
    ]:
        row2(ws1, r, label, val, unit); r += 1

    r += 1
    sec(ws1, r, 'REKOMENDOWANE ROZWIĄZANIA'); r += 1

    # Tabela produktów
    prod_headers = ['Produkt', 'Parametry', 'CAPEX (PLN)', 'Oszczędność/rok']
    for col, h in enumerate(prod_headers, 1):
        c = ws1.cell(row=r, column=col, value=h)
        c.font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
        c.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        c.border = thin; c.alignment = Alignment(horizontal='center', wrap_text=True)
    r += 1

    products = [
        ('Kontrakt ee', f'Produkt: {rek_ee.produkt}', '-', f'{rek_ee.oszczednosc_roczna_pln:,.0f}'),
        ('BESS', f'{rek_bess.pojemnosc_kwh:.0f} kWh / {rek_bess.moc_kw:.0f} kW', f'{rek_bess.capex_pln:,.0f}', f'{rek_bess.oszczednosc_calkowita_pln:,.0f}'),
    ]
    if rek_pv:
        products.append(('PV', f'{rek_pv.nowa_moc_kwp:.0f} kWp (nowa)', f'{rek_pv.capex_pln:,.0f}', f'{rek_pv.oszczednosc_roczna_pln:,.0f}'))
    if rek_dsr and rek_dsr.potencjal_redukcji_kw > 0:
        products.append(('DSR', f'{rek_dsr.potencjal_redukcji_kw:.0f} kW redukcji', f'{rek_dsr.koszt_wdrozenia_pln:,.0f}', f'{rek_dsr.przychod_roczny_pln:,.0f}'))
    if rek_kmb:
        products.append(('KMB', f'{rek_kmb.moc_potrzebna_kvar:.0f} kvar', f'{rek_kmb.capex_pln:,.0f}', f'{rek_kmb.oszczednosc_roczna_pln:,.0f}'))

    for prod_name, params, capex_str, oszcz_str in products:
        for col, val in enumerate([prod_name, params, capex_str, oszcz_str], 1):
            c = ws1.cell(row=r, column=col, value=val)
            c.font = val_font; c.border = thin
            if col >= 3:
                c.alignment = Alignment(horizontal='right')
        r += 1

    # TOTAL
    for col, val in enumerate(['RAZEM', '', f'{capex_total:,.0f}', f'{oszcz_total:,.0f}'], 1):
        c = ws1.cell(row=r, column=col, value=val)
        c.font = money_font; c.border = thin; c.fill = hi_fill
        if col >= 3:
            c.alignment = Alignment(horizontal='right')
    r += 2

    sec(ws1, r, 'KLUCZOWE WSKAŹNIKI'); r += 1
    okres_zw = capex_total / oszcz_total if oszcz_total > 0 else 99
    roczny_rach = dane.sredni_rachunek_ee_mies_pln * 12 + dane.sredni_rachunek_gaz_mies_pln * 12
    redukcja = oszcz_total / roczny_rach * 100 if roczny_rach > 0 else 0

    for label, val, unit, hi in [
        ('Łączny CAPEX', capex_total, 'PLN netto', False),
        ('Łączna oszczędność roczna', oszcz_total, 'PLN/rok', True),
        ('Prosty okres zwrotu', round(okres_zw, 1), 'lat', True),
        ('Redukcja kosztów energii', round(redukcja, 0), '%', True),
        ('Redukcja CO₂', round(dane.roczne_zuzycie_ee_kwh * 0.0003, 1), 'ton/rok', False),
    ]:
        row2(ws1, r, label, val, unit, bold=True, hi=hi); r += 1

    # ====== ARK 2: ENERGIA ELEKTRYCZNA ======
    ws2 = wb.create_sheet('Energia elektryczna')
    sw(ws2, [30, 15, 18, 20])

    r = 1
    hdr(ws2, r, 'REKOMENDACJA: KONTRAKT NA ENERGIĘ ELEKTRYCZNĄ'); r += 2

    sec(ws2, r, 'ANALIZA OBECNEGO KONTRAKTU'); r += 1
    for label, val, unit in [
        ('Obecny sprzedawca / taryfa', f'{dane.typ_umowy_ee}', ''),
        ('Koniec umowy', dane.data_konca_umowy_ee, ''),
        ('Obecna cena ee', dane.cena_ee_pln_kwh, 'PLN/kWh'),
        ('Roczne zużycie', dane.roczne_zuzycie_ee_kwh, 'kWh'),
        ('Roczny koszt ee (energia)', dane.roczne_zuzycie_ee_kwh * dane.cena_ee_pln_kwh, 'PLN'),
        ('Roczny koszt ee (pełny)', dane.sredni_rachunek_ee_mies_pln * 12, 'PLN'),
    ]:
        row2(ws2, r, label, val, unit); r += 1

    r += 1
    sec(ws2, r, 'PORÓWNANIE PRODUKTÓW'); r += 1

    cmp_headers = ['Produkt', 'Cena (PLN/kWh)', 'Roczny koszt ee', 'Oszczędność vs obecna']
    for col, h in enumerate(cmp_headers, 1):
        c = ws2.cell(row=r, column=col, value=h)
        c.font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
        c.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        c.border = thin; c.alignment = Alignment(horizontal='center')
    r += 1

    for prod, cena in [('FIX', rek_ee.cena_fix_pln_kwh), ('RDN (średnia)', rek_ee.cena_rdn_srednia_pln_kwh), ('MIX (50/50)', rek_ee.cena_mix_pln_kwh)]:
        roczny_koszt = cena * dane.roczne_zuzycie_ee_kwh
        oszcz = (dane.cena_ee_pln_kwh - cena) * dane.roczne_zuzycie_ee_kwh
        is_rec = prod.startswith(rek_ee.produkt)
        for col, val in enumerate([prod + (' ← REKOMENDACJA' if is_rec else ''), cena, f'{roczny_koszt:,.0f}', f'{oszcz:,.0f}'], 1):
            c = ws2.cell(row=r, column=col, value=val)
            c.font = money_font if is_rec else val_font
            c.border = thin
            if col >= 3:
                c.alignment = Alignment(horizontal='right')
            if is_rec:
                c.fill = hi_fill
        r += 1

    r += 1
    sec(ws2, r, 'UZASADNIENIE REKOMENDACJI'); r += 1
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    c = ws2.cell(row=r, column=1, value=rek_ee.uzasadnienie)
    c.font = val_font; c.alignment = Alignment(wrap_text=True)
    ws2.row_dimensions[r].height = 60

    r += 2
    sec(ws2, r, 'ANALIZA KOSZTÓW DYSTRYBUCJI'); r += 1
    koszt_dystr = dane.roczne_zuzycie_ee_kwh * dane.oplata_dystr_pln_kwh
    oszcz_dystr = koszt_dystr * 0.30  # 30% od oszczędności
    for label, val, unit in [
        ('Roczny koszt dystrybucji', koszt_dystr, 'PLN'),
        ('Potencjalna oszczędność (30%)', oszcz_dystr, 'PLN/rok'),
        ('Źródło oszczędności', 'Optymalizacja grupy taryfowej, mocy umownej, peak shaving', ''),
    ]:
        row2(ws2, r, label, val, unit); r += 1

    # ====== ARK 3: BESS ======
    ws3 = wb.create_sheet('Magazyn energii BESS')
    sw(ws3, [30, 15, 18, 20])

    r = 1
    hdr(ws3, r, 'REKOMENDACJA: MAGAZYN ENERGII (BESS)'); r += 2

    sec(ws3, r, 'PROPONOWANY SYSTEM'); r += 1
    for label, val, unit in [
        ('Pojemność', rek_bess.pojemnosc_kwh, 'kWh'),
        ('Moc', rek_bess.moc_kw, 'kW'),
        ('Technologia', 'LFP (LiFePO₄)', ''),
        ('Sprawność RTE', '90%', ''),
        ('Żywotność', '15 lat / 6000+ cykli', ''),
        ('Integracja', 'AC-coupled z istniejącym PV', ''),
        ('EMS', 'ENNO-EMS (ceny 15-min TGE, AI)', ''),
        ('Taryfa', 'Dynamiczna (15-min)', ''),
    ]:
        row2(ws3, r, label, val, unit); r += 1

    r += 1
    sec(ws3, r, 'CAPEX'); r += 1
    capex_bess_only = rek_bess.pojemnosc_kwh * 2000
    for label, val in [
        ('Magazyn energii (BESS)', capex_bess_only),
        ('Oprogramowanie EMS', 30000),
        ('Instalacja + zabezpieczenia', capex_bess_only * 0.10),
    ]:
        row2(ws3, r, label, val, 'PLN netto'); r += 1
    row2(ws3, r, 'CAPEX RAZEM', rek_bess.capex_pln, 'PLN netto', bold=True, hi=True); r += 2

    sec(ws3, r, 'OSZCZĘDNOŚCI ROCZNE (3 strumienie)'); r += 1
    for label, val in [
        ('1. Autokonsumpcja PV (nadwyżki → magazyn)', rek_bess.oszczednosc_autokonsumpcja_pln),
        ('2. Arbitraż cenowy (taryfa dynamiczna TGE)', rek_bess.oszczednosc_arbitraz_pln),
        ('3. Peak shaving (redukcja opłaty mocowej)', rek_bess.oszczednosc_peak_shaving_pln),
    ]:
        row2(ws3, r, label, val, 'PLN/rok', bold=True); r += 1
    row2(ws3, r, 'OSZCZĘDNOŚĆ ROCZNA RAZEM', rek_bess.oszczednosc_calkowita_pln, 'PLN/rok', bold=True, hi=True); r += 1
    row2(ws3, r, 'Okres zwrotu (prosty)', round(rek_bess.okres_zwrotu_lat, 1), 'lat', bold=True, hi=True); r += 2

    sec(ws3, r, 'STRATEGIA DZIAŁANIA'); r += 1
    strategie = [
        '10:00-15:00: PV → autokonsumpcja + nadwyżki do BESS. BESS kupuje tanią ee z TGE (ceny bliskie 0 / ujemne)',
        '17:00-21:00: BESS rozładowuje → zakład nie kupuje drogiej ee ze szczytu. Peak shaving opłaty mocowej',
        '22:00-06:00: BESS ładuje po niskich cenach nocnych z TGE',
        'EMS automatycznie optymalizuje na podstawie cen 15-min TGE, prognoz pogody i profilu zużycia',
    ]
    for s in strategie:
        ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = ws3.cell(row=r, column=1, value=f'  →  {s}')
        c.font = val_font; c.alignment = Alignment(wrap_text=True)
        ws3.row_dimensions[r].height = 30
        r += 1

    # ====== ARK 4: PV + DSR + KMB ======
    ws4 = wb.create_sheet('PV + DSR + KMB')
    sw(ws4, [30, 15, 18, 20])

    r = 1
    hdr(ws4, r, 'DODATKOWE REKOMENDACJE'); r += 2

    # PV
    if rek_pv:
        sec(ws4, r, 'FOTOWOLTAIKA (PV) – nowa instalacja'); r += 1
        for label, val, unit in [
            ('Rekomendowana moc', rek_pv.nowa_moc_kwp, 'kWp'),
            ('Roczna produkcja', rek_pv.roczna_produkcja_kwh, 'kWh/rok'),
            ('CAPEX', rek_pv.capex_pln, 'PLN netto'),
            ('Oszczędność roczna', rek_pv.oszczednosc_roczna_pln, 'PLN/rok'),
            ('Okres zwrotu', round(rek_pv.okres_zwrotu_lat, 1), 'lat'),
            ('Autokonsumpcja (bez BESS)', rek_pv.autokonsumpcja_procent, '%'),
            ('Partner realizacji', 'SUN HELP', ''),
        ]:
            row2(ws4, r, label, val, unit, bold=(label in ['Oszczędność roczna', 'Okres zwrotu'])); r += 1
        r += 1
    else:
        sec(ws4, r, 'FOTOWOLTAIKA (PV)'); r += 1
        ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws4.cell(row=r, column=1, value='Istniejąca instalacja PV jest wystarczająca lub brak miejsca na dodatkowe panele.').font = val_font
        r += 2

    # DSR
    sec(ws4, r, 'DSR (Demand Side Response)'); r += 1
    if rek_dsr and rek_dsr.potencjal_redukcji_kw > 0:
        for label, val, unit in [
            ('Potencjał redukcji', rek_dsr.potencjal_redukcji_kw, 'kW'),
            ('Przychód roczny', rek_dsr.przychod_roczny_pln, 'PLN/rok'),
            ('Koszt wdrożenia', rek_dsr.koszt_wdrozenia_pln, 'PLN'),
            ('Partner realizacji', 'ALIANS', ''),
        ]:
            row2(ws4, r, label, val, unit, bold=(label == 'Przychód roczny')); r += 1
        ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws4.cell(row=r, column=1, value=rek_dsr.uzasadnienie).font = note_font
        ws4.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
        ws4.row_dimensions[r].height = 50
        r += 2
    else:
        ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws4.cell(row=r, column=1, value=rek_dsr.uzasadnienie if rek_dsr else 'Brak potencjału DSR.').font = val_font
        r += 2

    # KMB
    sec(ws4, r, 'KMB (Kompensacja Mocy Biernej)'); r += 1
    if rek_kmb:
        for label, val, unit in [
            ('Potrzebna moc kompensacji', rek_kmb.moc_potrzebna_kvar, 'kvar'),
            ('CAPEX', rek_kmb.capex_pln, 'PLN netto'),
            ('Oszczędność roczna', rek_kmb.oszczednosc_roczna_pln, 'PLN/rok'),
            ('Okres zwrotu', round(rek_kmb.okres_zwrotu_lat, 1), 'lat'),
        ]:
            row2(ws4, r, label, val, unit, bold=(label == 'Oszczędność roczna')); r += 1
    else:
        ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        msg = 'Kompensacja mocy biernej już zainstalowana.' if dane.ma_kmb else 'cos(φ) ≥ 0.95 – kompensacja nie jest potrzebna.'
        ws4.cell(row=r, column=1, value=msg).font = val_font
        r += 1

    # ====== ARK 5: FINANSOWANIE ======
    ws5 = wb.create_sheet('Opcje finansowania')
    sw(ws5, [28, 35, 15, 15, 15, 30])

    r = 1
    ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws5.cell(row=r, column=1, value=f'OPCJE FINANSOWANIA – łączny CAPEX: {capex_total:,.0f} PLN netto')
    c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal='center', vertical='center')
    ws5.row_dimensions[r].height = 35
    r += 2

    fin_headers = ['Model', 'Opis', 'Wkład własny', 'Rata/mies.', 'Okres', 'Uwagi']
    for col, h in enumerate(fin_headers, 1):
        c = ws5.cell(row=r, column=col, value=h)
        c.font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
        c.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        c.border = thin; c.alignment = Alignment(horizontal='center', wrap_text=True)
    ws5.row_dimensions[r].height = 30
    r += 1

    for opcja in opcje_fin:
        vals = [
            opcja.nazwa,
            opcja.opis,
            f'{opcja.wklad_wlasny_procent:.0f}%',
            f'{opcja.rata_miesieczna_pln:,.0f} PLN' if opcja.rata_miesieczna_pln > 0 else '-',
            f'{opcja.okres_mies} mies.' if opcja.okres_mies > 0 else '-',
            opcja.uwagi,
        ]
        for col, val in enumerate(vals, 1):
            c = ws5.cell(row=r, column=col, value=val)
            c.font = val_font; c.border = thin
            c.alignment = Alignment(wrap_text=True, vertical='top')
        ws5.row_dimensions[r].height = 55
        r += 1

    r += 2
    ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws5.cell(row=r, column=1, value='PORÓWNANIE KOSZTÓW FINANSOWANIA')
    c.font = sec_font; c.fill = sec_fill
    r += 1

    cmp_h = ['Model', 'CAPEX efektywny', 'Rata/mies.', 'Koszt całkowity', 'Korzyść podat./rok', 'Efektywny koszt netto*']
    for col, h in enumerate(cmp_h, 1):
        c = ws5.cell(row=r, column=col, value=h)
        c.font = Font(name='Calibri', bold=True, size=9, color='FFFFFF')
        c.fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        c.border = thin; c.alignment = Alignment(horizontal='center', wrap_text=True)
    r += 1

    for opcja in opcje_fin:
        capex_ef = capex_total * opcja.wklad_wlasny_procent / 100
        koszt_netto = opcja.koszt_calkowity_pln - opcja.korzysc_podatkowa_roczna_pln * (opcja.okres_mies / 12 if opcja.okres_mies > 0 else 10)
        vals = [
            opcja.nazwa,
            f'{capex_ef:,.0f} PLN',
            f'{opcja.rata_miesieczna_pln:,.0f} PLN' if opcja.rata_miesieczna_pln > 0 else '-',
            f'{opcja.koszt_calkowity_pln:,.0f} PLN',
            f'{opcja.korzysc_podatkowa_roczna_pln:,.0f} PLN',
            f'{max(0, koszt_netto):,.0f} PLN',
        ]
        for col, val in enumerate(vals, 1):
            c = ws5.cell(row=r, column=col, value=val)
            c.font = val_font; c.border = thin
            c.alignment = Alignment(horizontal='right' if col > 1 else 'left')
        r += 1

    r += 1
    ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws5.cell(row=r, column=1, value='* Efektywny koszt netto = koszt całkowity - korzyści podatkowe w okresie finansowania. Nie uwzględnia oszczędności z eksploatacji.').font = note_font

    # ====== ARK 6: ANALIZA 10-LETNIA ======
    ws6 = wb.create_sheet('Analiza 10-letnia')
    sw(ws6, [8, 18, 18, 18, 18, 18])

    r = 1
    ws6.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws6.cell(row=r, column=1, value='PROJEKCJA FINANSOWA – 10 LAT')
    c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal='center', vertical='center')
    ws6.row_dimensions[r].height = 35
    r += 2

    yr_headers = ['Rok', 'Oszczędność\nbazowa (PLN)', 'Wzrost cen\nee (+3%/rok)', 'Oszczędność\nskorag. (PLN)', 'Cash flow\nskumulowany', 'Oszczędność\nvs leasing']
    for col, h in enumerate(yr_headers, 1):
        c = ws6.cell(row=r, column=col, value=h)
        c.font = Font(name='Calibri', bold=True, size=9, color='FFFFFF')
        c.fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        c.border = thin; c.alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')
    ws6.row_dimensions[r].height = 40
    r += 1

    # Rok 0
    skum = -capex_total
    rata_leasing_best = opcje_fin[1].rata_miesieczna_pln  # leasing operacyjny
    for col, val in enumerate([0, 0, '0%', f'{-capex_total:,.0f}', f'{skum:,.0f}', '-'], 1):
        c = ws6.cell(row=r, column=col, value=val)
        c.font = val_font; c.border = thin; c.alignment = Alignment(horizontal='right')
    r += 1

    for rok in range(1, 11):
        wzrost_cen = 1.03 ** rok  # 3% wzrost cen ee rocznie
        degradacja = (1 - 0.02) ** rok  # 2% degradacja BESS
        oszcz_skorygowana = oszcz_total * wzrost_cen * degradacja
        skum += oszcz_skorygowana
        leasing_oszcz = oszcz_skorygowana - rata_leasing_best * 12 if rok <= 7 else oszcz_skorygowana

        for col, val in enumerate([
            rok,
            f'{oszcz_total:,.0f}',
            f'+{(wzrost_cen - 1) * 100:.0f}%',
            f'{oszcz_skorygowana:,.0f}',
            f'{skum:,.0f}',
            f'{leasing_oszcz:,.0f}',
        ], 1):
            c = ws6.cell(row=r, column=col, value=val)
            c.font = val_font; c.border = thin
            c.alignment = Alignment(horizontal='right')
            if col == 5 and skum >= 0:
                c.fill = hi_fill; c.font = money_font
        r += 1

    # ====== ARK 7: KORZYŚCI ======
    ws7 = wb.create_sheet('Korzyści')
    sw(ws7, [55, 25])

    r = 1
    ws7.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    c = ws7.cell(row=r, column=1, value='PEŁNA LISTA KORZYŚCI')
    c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal='center', vertical='center')
    ws7.row_dimensions[r].height = 35
    r += 2

    korzysci_sections = [
        ('OSZCZĘDNOŚCI FINANSOWE', [
            f'Łączna oszczędność roczna: {oszcz_total:,.0f} PLN',
            f'Redukcja rachunku za energię o {redukcja:.0f}%',
            f'Okres zwrotu inwestycji: {okres_zw:.1f} lat',
            'Ochrona przed wzrostem cen energii (hedging)',
            f'Prognozowana oszczędność w 10 lat: {oszcz_total * 10.5:,.0f} PLN (z uwzględnieniem wzrostu cen)',
        ]),
        ('OPTYMALIZACJA KONTRAKTU EE', [
            f'Rekomendowany produkt: {rek_ee.produkt} – oszczędność {rek_ee.oszczednosc_roczna_pln:,.0f} PLN/rok',
            'Możliwość korzystania z cen 15-minutowych TGE (taryfa dynamiczna)',
            'Zakup energii po cenach bliskich zeru lub ujemnych (10:00-15:00)',
            'Optymalizacja kosztów dystrybucji (30% potencjał oszczędności)',
        ]),
        ('MAGAZYN ENERGII (BESS)', [
            f'Autokonsumpcja PV: nadwyżki do magazynu – {rek_bess.oszczednosc_autokonsumpcja_pln:,.0f} PLN/rok',
            f'Arbitraż cenowy (kupno tanio / sprzedaż drogo) – {rek_bess.oszczednosc_arbitraz_pln:,.0f} PLN/rok',
            f'Peak shaving (redukcja opłaty mocowej) – {rek_bess.oszczednosc_peak_shaving_pln:,.0f} PLN/rok',
            'Możliwość świadczenia usług pomocniczych (aFRR, mFRR) – dodatkowe przychody',
            'Uczestnictwo w rynku mocy (capacity market)',
            'Backup zasilania – ochrona przed przerwami w dostawie',
        ]),
        ('NIEZALEŻNOŚĆ ENERGETYCZNA', [
            'Wzrost autokonsumpcji PV z magazynem',
            'Mniejsza zależność od wahań cen rynkowych',
            'Przygotowanie na regulacje klimatyczne (ETS2, CBAM)',
            'Możliwość pracy wyspowej (island mode) z BESS',
        ]),
        ('ESG I WIZERUNEK', [
            f'Redukcja emisji CO₂: {dane.roczne_zuzycie_ee_kwh * 0.0003:.1f} ton/rok',
            'Gwarancje pochodzenia energii (zielona energia)' if dane.potrzebuje_go else 'Możliwość dokupienia gwarancji pochodzenia (GO)',
            'Raportowanie ESG – niższy ślad węglowy',
            'Spełnienie wymagań klientów korporacyjnych (supply chain ESG)',
        ]),
        ('ELASTYCZNE FINANSOWANIE', [
            'Leasing operacyjny: 0% wkładu, cała rata w KUP',
            'Kredyt ekologiczny BGK: premia do 80% kosztów',
            'Model ESCO: zero CAPEX, spłata z oszczędności',
            'On-site PPA: stała cena energii 10-15 lat',
        ]),
    ]

    for section_title, items in korzysci_sections:
        ws7.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        c = ws7.cell(row=r, column=1, value=section_title)
        c.font = sec_font; c.fill = sec_fill
        r += 1
        for item in items:
            ws7.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            c = ws7.cell(row=r, column=1, value=f'  ✓  {item}')
            c.font = val_font
            r += 1
        r += 1

    # Zapisz
    wb.save(output_path)
    print(f'\nOferta zapisana: {output_path}')


def generuj_oferte_bytes(dane: DaneKlienta) -> bytes:
    """Generuje ofertę XLSX i zwraca jako bytes (do st.download_button)."""
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        tmp_path = tmp.name
    try:
        generuj_oferte(dane, tmp_path)
        with open(tmp_path, 'rb') as f:
            return f.read()
    finally:
        os.unlink(tmp_path)


# ============================================================
# INTERACTIVE INPUT
# ============================================================

def ask_float(prompt, default=None):
    while True:
        try:
            suffix = f' [{default}]' if default is not None else ''
            val = input(f'  {prompt}{suffix}: ').strip()
            if not val and default is not None:
                return float(default)
            return float(val.replace(',', '.').replace(' ', ''))
        except ValueError:
            print('  Podaj liczbę.')

def ask_str(prompt, default=''):
    suffix = f' [{default}]' if default else ''
    val = input(f'  {prompt}{suffix}: ').strip()
    return val if val else default

def ask_bool(prompt, default=False):
    suffix = f' [{"T" if default else "N"}]'
    val = input(f'  {prompt} (T/N){suffix}: ').strip().upper()
    if not val:
        return default
    return val.startswith('T') or val.startswith('Y')


def pobierz_dane() -> DaneKlienta:
    print('\n' + '=' * 65)
    print('  KALKULATOR OFERTOWY: ee + gaz + PV + BESS + DSR + KMB')
    print('=' * 65)

    print('\n--- DANE FIRMY ---')
    nazwa = ask_str('Nazwa firmy', 'Firma Sp. z o.o.')
    nip = ask_str('NIP', '0000000000')
    branza = ask_str('Branża', 'Produkcja')
    dni = ask_str('Dni pracy (np. Pn-Pt)', 'Pn-Pt')
    godz = ask_str('Godziny pracy (np. 6:00-22:00)', '6:00-22:00 (2 zmiany)')

    print('\n--- ENERGIA ELEKTRYCZNA ---')
    zuzycie = ask_float('Roczne zużycie ee (kWh/rok)')
    moc_um = ask_float('Moc umowna (kW)')
    moc_przyl = ask_float('Moc przyłączeniowa (kW)', moc_um)
    grupa = ask_str('Grupa taryfowa', 'C22a')
    osd = ask_str('OSD', 'Tauron')
    rachunek_ee = ask_float('Średni rachunek ee (PLN/mies.)')
    cena_ee = ask_float('Cena energii z umowy (PLN/kWh)', 0.65)
    dystr = ask_float('Opłata dystrybucyjna (PLN/kWh)', 0.25)
    mocowa = ask_float('Stawka opłaty mocowej (PLN/MWh)', 219.40)
    kat = ask_str('Kategoria mocowa (K1-K4)', 'K3').upper()
    koniec_ee = ask_str('Data końca umowy ee', '2026-12-31')
    typ_ee = ask_str('Typ obecnej umowy (FIX/RDN/MIX)', 'FIX')

    print('\n--- GAZ ---')
    zuzycie_gaz = ask_float('Roczne zużycie gazu (kWh/rok)', 0)
    rachunek_gaz = ask_float('Średni rachunek gaz (PLN/mies.)', 0)
    cena_gaz = ask_float('Cena gazu (PLN/kWh)', 0.25)
    koniec_gaz = ask_str('Data końca umowy gaz', '2026-12-31')

    print('\n--- INSTALACJA PV ---')
    ma_pv = ask_bool('Czy jest instalacja PV?', False)
    moc_pv = ask_float('Moc PV (kWp)', 0) if ma_pv else 0
    prod_pv = ask_float('Roczna produkcja PV (kWh)', moc_pv * 1050) if ma_pv else 0
    autokons = ask_float('Autokonsumpcja PV (%)', 35) if ma_pv else 0

    print('\n--- INFRASTRUKTURA ---')
    ma_kmb = ask_bool('Czy jest kompensacja mocy biernej (KMB)?', False)
    moc_bierna = ask_float('Moc kompensacji (kvar)', 0) if ma_kmb else 0
    ma_agregat = ask_bool('Czy jest agregat prądotwórczy?', False)
    go = ask_bool('Czy potrzebne gwarancje pochodzenia?', False)
    dach = ask_float('Wolna powierzchnia na PV (m²)', 500)
    cos_phi = ask_float('Współczynnik cos(φ)', 0.85)

    return DaneKlienta(
        nazwa_firmy=nazwa, nip=nip, branza=branza,
        dni_pracy=dni, godziny_pracy=godz,
        roczne_zuzycie_ee_kwh=zuzycie, moc_umowna_kw=moc_um,
        moc_przylaczeniowa_kw=moc_przyl, grupa_taryfowa=grupa,
        osd=osd, sredni_rachunek_ee_mies_pln=rachunek_ee,
        cena_ee_pln_kwh=cena_ee, oplata_dystr_pln_kwh=dystr,
        oplata_mocowa_pln_mwh=mocowa, kategoria_mocowa=kat,
        data_konca_umowy_ee=koniec_ee, typ_umowy_ee=typ_ee,
        roczne_zuzycie_gaz_kwh=zuzycie_gaz,
        sredni_rachunek_gaz_mies_pln=rachunek_gaz,
        cena_gaz_pln_kwh=cena_gaz, data_konca_umowy_gaz=koniec_gaz,
        ma_pv=ma_pv, moc_pv_kwp=moc_pv,
        roczna_produkcja_pv_kwh=prod_pv,
        autokonsumpcja_pv_procent=autokons,
        ma_kmb=ma_kmb, moc_bierna_kvar=moc_bierna,
        ma_agregat=ma_agregat, potrzebuje_go=go,
        powierzchnia_dachu_m2=dach, wspolczynnik_cos_phi=cos_phi,
    )


def dane_demo() -> DaneKlienta:
    return DaneKlienta(
        nazwa_firmy='Przykładowy Zakład Produkcyjny Sp. z o.o.',
        nip='1234567890', branza='Produkcja metalowa',
        dni_pracy='Pn-Pt', godziny_pracy='6:00-22:00 (2 zmiany)',
        roczne_zuzycie_ee_kwh=800_000, moc_umowna_kw=350,
        moc_przylaczeniowa_kw=400, grupa_taryfowa='C22a',
        osd='Tauron', sredni_rachunek_ee_mies_pln=55_000,
        cena_ee_pln_kwh=0.68, oplata_dystr_pln_kwh=0.27,
        oplata_mocowa_pln_mwh=219.40, kategoria_mocowa='K3',
        data_konca_umowy_ee='2026-09-30', typ_umowy_ee='FIX',
        roczne_zuzycie_gaz_kwh=200_000,
        sredni_rachunek_gaz_mies_pln=8_000,
        cena_gaz_pln_kwh=0.28, data_konca_umowy_gaz='2026-12-31',
        ma_pv=True, moc_pv_kwp=200,
        roczna_produkcja_pv_kwh=210_000,
        autokonsumpcja_pv_procent=35,
        ma_kmb=False, moc_bierna_kvar=0,
        ma_agregat=False, potrzebuje_go=True,
        powierzchnia_dachu_m2=800,
        wspolczynnik_cos_phi=0.85,
    )


def main():
    demo = '--demo' in sys.argv
    dane = dane_demo() if demo else pobierz_dane()

    if demo:
        print('\n  Tryb DEMO – dane przykładowe\n')

    output = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                          f'Oferta_{dane.nazwa_firmy.replace(" ", "_")[:30]}.xlsx')

    generuj_oferte(dane, output)

    print(f'\n  Oferta wygenerowana: {output}\n')


if __name__ == '__main__':
    main()
